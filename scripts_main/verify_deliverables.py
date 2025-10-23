"""
成果物の自動検証スクリプト

検証内容:
- ファイルの存在確認
- CSV/Excelの列数・行数確認
- Excelシートの存在確認
- データ整合性チェック
"""
import os
import pandas as pd
import openpyxl
from datetime import datetime

# パス設定
PROJECT_ROOT = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
DELIVERABLES_DIR = os.path.join(PROJECT_ROOT, '成果物_20251023')
FOR_CLAUDE_DIR = os.path.join(PROJECT_ROOT, 'for_claude')


def check_file_exists(file_path, file_description):
    """ファイルの存在確認"""
    if os.path.exists(file_path):
        file_size_kb = os.path.getsize(file_path) / 1024
        print(f"OK: {file_description}")
        print(f"    パス: {file_path}")
        print(f"    サイズ: {file_size_kb:.2f} KB")
        return True
    else:
        print(f"NG: {file_description} が見つかりません")
        print(f"    パス: {file_path}")
        return False


def verify_csv_file(file_path):
    """CSVファイルの検証"""
    print()
    print("-" * 80)
    print("CSV検証")
    print("-" * 80)

    if not check_file_exists(file_path, "インポート用CSV"):
        return False

    try:
        df = pd.read_csv(file_path, encoding='utf-8-sig')

        print(f"    行数: {len(df)}行")
        print(f"    列数: {len(df.columns)}列")

        # 列数チェック（38列期待）
        if len(df.columns) == 38:
            print(f"    OK: 列数が正しい（38列）")
        else:
            print(f"    NG: 列数が不正（期待: 38列、実際: {len(df.columns)}列）")
            return False

        # 必須列の存在確認
        required_columns = ['ja', 'en', 'fil-PH', 'zh', 'th', 'vi', 'my', 'id', 'km']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            print(f"    NG: 必須列が不足: {missing_columns}")
            return False
        else:
            print(f"    OK: 必須列が存在")

        # データの存在確認
        if len(df) > 0:
            print(f"    OK: データが存在（{len(df)}行）")
        else:
            print(f"    NG: データが空")
            return False

        print()
        print("CSV検証: 合格")
        return True

    except Exception as e:
        print(f"    エラー: {str(e)}")
        return False


def verify_excel_file(file_path):
    """Excelファイルの検証"""
    print()
    print("-" * 80)
    print("Excel検証")
    print("-" * 80)

    if not check_file_exists(file_path, "比較Excel"):
        return False

    try:
        wb = openpyxl.load_workbook(file_path)

        print(f"    シート数: {len(wb.sheetnames)}")
        print(f"    シート名: {', '.join(wb.sheetnames)}")

        # 必須シートの確認
        required_sheets = ['翻訳', '再翻訳', '比較']
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]

        if missing_sheets:
            print(f"    NG: 必須シートが不足: {missing_sheets}")
            return False
        else:
            print(f"    OK: 必須シートが存在")

        # 翻訳シート検証
        print()
        print("  【翻訳シート】")
        df_translation = pd.read_excel(file_path, sheet_name='翻訳', header=0)
        print(f"    行数: {len(df_translation)}行")
        print(f"    列数: {len(df_translation.columns)}列")

        if len(df_translation.columns) == 38:
            print(f"    OK: 列数が正しい（38列）")
        else:
            print(f"    NG: 列数が不正（期待: 38列、実際: {len(df_translation.columns)}列）")
            return False

        # 再翻訳シート検証
        print()
        print("  【再翻訳シート】")
        df_retranslation = pd.read_excel(file_path, sheet_name='再翻訳', header=0)
        print(f"    行数: {len(df_retranslation)}行")
        print(f"    列数: {len(df_retranslation.columns)}列")

        if len(df_retranslation.columns) == 38:
            print(f"    OK: 列数が正しい（38列）")
        else:
            print(f"    NG: 列数が不正（期待: 38列、実際: {len(df_retranslation.columns)}列）")
            return False

        # 行数の一致確認
        if len(df_translation) == len(df_retranslation):
            print(f"    OK: 翻訳と再翻訳の行数が一致（{len(df_translation)}行）")
        else:
            print(f"    NG: 翻訳と再翻訳の行数が不一致")
            print(f"        翻訳: {len(df_translation)}行")
            print(f"        再翻訳: {len(df_retranslation)}行")
            return False

        # 比較シート検証
        print()
        print("  【比較シート】")
        df_comparison = pd.read_excel(file_path, sheet_name='比較', header=0)
        print(f"    行数: {len(df_comparison)}行")
        print(f"    列数: {len(df_comparison.columns)}列")

        # 必須列の確認
        required_comparison_columns = ['アドレス', '言語', '単語', '再翻訳', '翻訳']
        missing_comparison_columns = [col for col in required_comparison_columns if col not in df_comparison.columns]

        if missing_comparison_columns:
            print(f"    NG: 必須列が不足: {missing_comparison_columns}")
            return False
        else:
            print(f"    OK: 必須列が存在")

        # 類似度列の確認
        if '類似度_difflib' in df_comparison.columns:
            print(f"    OK: 類似度_difflib列が存在")

            # 類似度の統計
            if '分類' in df_comparison.columns:
                classification_counts = df_comparison['分類'].value_counts()
                print()
                print("    【類似度分類】")
                for category, count in classification_counts.items():
                    percentage = (count / len(df_comparison)) * 100
                    print(f"      {category}: {count}件 ({percentage:.1f}%)")
        else:
            print(f"    警告: 類似度_difflib列が見つかりません")

        print()
        print("Excel検証: 合格")
        return True

    except Exception as e:
        print(f"    エラー: {str(e)}")
        return False


def verify_data_consistency(csv_path, excel_path):
    """データ整合性の検証"""
    print()
    print("-" * 80)
    print("データ整合性検証")
    print("-" * 80)

    try:
        # CSVとExcelの行数比較
        df_csv = pd.read_csv(csv_path, encoding='utf-8-sig')
        df_excel_translation = pd.read_excel(excel_path, sheet_name='翻訳', header=0)

        if len(df_csv) == len(df_excel_translation):
            print(f"OK: CSVとExcelの行数が一致（{len(df_csv)}行）")
        else:
            print(f"NG: CSVとExcelの行数が不一致")
            print(f"    CSV: {len(df_csv)}行")
            print(f"    Excel: {len(df_excel_translation)}行")
            return False

        # 日本語列の一致確認
        csv_ja = df_csv['ja'].tolist()
        excel_ja = df_excel_translation['ja'].tolist()

        mismatch_count = 0
        for i, (csv_val, excel_val) in enumerate(zip(csv_ja, excel_ja)):
            if str(csv_val).strip() != str(excel_val).strip():
                mismatch_count += 1
                if mismatch_count <= 5:  # 最初の5件のみ表示
                    print(f"    不一致 (行{i+2}): CSV='{csv_val}' vs Excel='{excel_val}'")

        if mismatch_count == 0:
            print(f"OK: 日本語列（ja）が完全に一致")
        else:
            print(f"NG: 日本語列（ja）に{mismatch_count}件の不一致")
            return False

        print()
        print("データ整合性検証: 合格")
        return True

    except Exception as e:
        print(f"エラー: {str(e)}")
        return False


def main():
    """メイン処理"""
    print("=" * 80)
    print("成果物検証")
    print("=" * 80)
    print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # ファイルパス
    csv_file = os.path.join(DELIVERABLES_DIR, '01_全言語統合_テンプレート_インポート用.csv')
    excel_file = os.path.join(DELIVERABLES_DIR, '02_逆翻訳_検証結果.xlsx')

    # 検証実行
    results = []

    # CSV検証
    csv_result = verify_csv_file(csv_file)
    results.append(("CSV検証", csv_result))

    # Excel検証
    excel_result = verify_excel_file(excel_file)
    results.append(("Excel検証", excel_result))

    # データ整合性検証
    if csv_result and excel_result:
        consistency_result = verify_data_consistency(csv_file, excel_file)
        results.append(("データ整合性検証", consistency_result))
    else:
        print()
        print("-" * 80)
        print("データ整合性検証: スキップ（前提条件未達）")
        print("-" * 80)
        results.append(("データ整合性検証", False))

    # 結果サマリー
    print()
    print("=" * 80)
    print("検証結果サマリー")
    print("=" * 80)
    print()

    all_passed = True
    for test_name, result in results:
        status = "合格" if result else "不合格"
        symbol = "OK" if result else "NG"
        print(f"{symbol}: {test_name} ... {status}")
        if not result:
            all_passed = False

    print()
    print("=" * 80)
    if all_passed:
        print("総合判定: 合格")
        print()
        print("すべての検証に合格しました。")
        print("成果物は正しく生成されています。")
    else:
        print("総合判定: 不合格")
        print()
        print("一部の検証に失敗しました。")
        print("上記のエラーを確認して、再度処理を実行してください。")
    print("=" * 80)
    print()

    # 結果をファイルに保存
    output_file = os.path.join(FOR_CLAUDE_DIR, 'verification_result.txt')
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("成果物検証結果\n")
        f.write("=" * 80 + "\n")
        f.write(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        for test_name, result in results:
            status = "合格" if result else "不合格"
            f.write(f"{test_name}: {status}\n")

        f.write("\n")
        if all_passed:
            f.write("総合判定: 合格\n")
        else:
            f.write("総合判定: 不合格\n")

    print(f"検証結果を保存しました: {output_file}")
    print()


if __name__ == "__main__":
    main()
