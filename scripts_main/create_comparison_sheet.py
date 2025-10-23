"""
翻訳と再翻訳を比較して差異を抽出
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

project_root = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
deliverables_dir = os.path.join(project_root, '成果物_20251023')
output_dir = os.path.join(project_root, 'output')

print("=" * 80)
print("翻訳と再翻訳の比較シート作成")
print("=" * 80)
print()

# Excelファイル読み込み
excel_file = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

if not os.path.exists(excel_file):
    excel_file = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')

print(f"読み込み: {excel_file}")
print()

# シート読み込み
df_translation = pd.read_excel(excel_file, sheet_name='翻訳', header=0)
df_retranslation = pd.read_excel(excel_file, sheet_name='再翻訳', header=0)

print(f"翻訳シート: {len(df_translation)}行 x {len(df_translation.columns)}列")
print(f"再翻訳シート: {len(df_retranslation)}行 x {len(df_retranslation.columns)}列")
print()

# 比較データを格納するリスト
comparison_data = []

# 列名（日本語以外）
language_columns = [col for col in df_translation.columns if col != 'ja']

print("比較中...")
print()

# 各行を比較
for row_idx in range(len(df_translation)):
    japanese_word = df_translation.loc[row_idx, 'ja']

    # 各言語列を比較
    for col_idx, lang in enumerate(language_columns):
        # 翻訳と再翻訳の値を取得
        translation_val = df_translation.loc[row_idx, lang]
        retranslation_val = df_retranslation.loc[row_idx, lang]

        # NaNチェック
        if pd.isna(translation_val):
            translation_val = ''
        if pd.isna(retranslation_val):
            retranslation_val = ''

        # 文字列に変換
        translation_str = str(translation_val).strip()
        retranslation_str = str(retranslation_val).strip()

        # 差異がある場合
        if translation_str != retranslation_str:
            # セルアドレスを計算（Aは日本語、Bから各言語）
            # 列: A=ja, B=en, C=fil-PH, ...
            col_letter = chr(66 + col_idx)  # B, C, D, ...
            cell_address = f"{col_letter}{row_idx + 2}"  # +2 (ヘッダー1行 + 0-indexed)

            # 単語と再翻訳の一致判定
            match_result = japanese_word.strip() == retranslation_str.strip()

            comparison_data.append({
                'アドレス': cell_address,
                '言語': lang,
                '単語': japanese_word,
                '再翻訳': retranslation_str,
                '翻訳': translation_str,
                '一致': match_result
            })

print(f"差異検出: {len(comparison_data)}件")
print()

# 比較データをDataFrameに変換
df_comparison = pd.DataFrame(comparison_data)

if len(df_comparison) == 0:
    print("差異が見つかりませんでした。")
    print()
else:
    # 最初の10件を表示
    print("差異の例（最初の10件）:")
    print("-" * 80)
    try:
        for idx, row in df_comparison.head(10).iterrows():
            # エンコーディングエラー回避のため、シンプルな出力に
            print(f"アドレス: {row['アドレス']}, 言語: {row['言語']}, 一致: {row['一致']}")
            print(f"  単語: {row['単語']}")
            print(f"  再翻訳: {row['再翻訳']}")
            print(f"  翻訳: {row['翻訳']}")
            print()
    except UnicodeEncodeError:
        print("(差異の詳細表示をスキップしました)")
    print()

    # Excelに出力
    output_excel = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

    print("比較シートを追加中...")

    # 既存のExcelを開く
    wb = openpyxl.load_workbook(output_excel)

    # 既存の比較シートを削除（存在する場合）
    if '比較' in wb.sheetnames:
        wb.remove(wb['比較'])

    # 新しいシートを作成（最後の位置に追加）
    ws = wb.create_sheet('比較')

    # ヘッダー設定
    headers = ['アドレス', '言語', '単語', '再翻訳', '翻訳', '一致']
    ws.append(headers)

    # ヘッダーのスタイル設定（黄色背景）
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # データを追加
    for idx, row in df_comparison.iterrows():
        ws.append([
            row['アドレス'],
            row['言語'],
            row['単語'],
            row['再翻訳'],
            row['翻訳'],
            row['一致']
        ])

    # 列幅を調整
    ws.column_dimensions['A'].width = 10  # アドレス
    ws.column_dimensions['B'].width = 10  # 言語
    ws.column_dimensions['C'].width = 30  # 単語
    ws.column_dimensions['D'].width = 40  # 再翻訳
    ws.column_dimensions['E'].width = 40  # 翻訳
    ws.column_dimensions['F'].width = 10  # 一致

    # 保存
    wb.save(output_excel)

    print(f"保存: {output_excel}")
    print()

    # outputフォルダにもコピー
    output_excel_copy = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')
    wb.save(output_excel_copy)
    print(f"コピー保存: {output_excel_copy}")
    print()

print("=" * 80)
print("完了")
print("=" * 80)
print()
print(f"シート構成:")
print(f"  - 比較: {len(df_comparison)}件の差異")
print(f"  - 翻訳: {len(df_translation)}行（元の翻訳データ）")
print(f"  - 再翻訳: {len(df_retranslation)}行（日本語逆翻訳）")
print()
