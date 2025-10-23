"""
38列に拡張されたExcelファイルを検証
"""
import pandas as pd
import openpyxl
import os

project_root = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
deliverables_dir = os.path.join(project_root, '成果物_20251023')

excel_file = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

print("=" * 80)
print("38列Excelファイルの検証")
print("=" * 80)
print()

# Excelファイルを開く
wb = openpyxl.load_workbook(excel_file)

print(f"シート一覧: {wb.sheetnames}")
print()

# 各シートの詳細
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"[{sheet_name}]")
    print(f"  行数: {ws.max_row}行")
    print(f"  列数: {ws.max_column}列")

    # ヘッダー行を表示
    if ws.max_row > 0:
        headers = [cell.value for cell in ws[1]]
        print(f"  ヘッダー（最初の10列）: {headers[:10]}")
        print(f"  ヘッダー（最後の5列）: {headers[-5:]}")

    print()

# pandasで詳細確認
print("=" * 80)
print("データ内容確認")
print("=" * 80)
print()

# 翻訳シート
df_translation = pd.read_excel(excel_file, sheet_name='翻訳', header=0)
print("[翻訳シート]")
print(f"  形状: {df_translation.shape}")
print(f"  列名: {list(df_translation.columns[:5])} ... {list(df_translation.columns[-5:])}")
print()

# データがある列をカウント
non_empty_cols = []
for col in df_translation.columns:
    non_empty_count = df_translation[col].notna().sum()
    if non_empty_count > 0:
        non_empty_cols.append((col, non_empty_count))

print(f"データがある列: {len(non_empty_cols)}列")
for col, count in non_empty_cols[:10]:
    print(f"  {col}: {count}行")
print()

# 再翻訳シート
df_retranslation = pd.read_excel(excel_file, sheet_name='再翻訳', header=0)
print("[再翻訳シート]")
print(f"  形状: {df_retranslation.shape}")
print(f"  列名: {list(df_retranslation.columns[:5])} ... {list(df_retranslation.columns[-5:])}")
print()

# 比較シート
if '比較' in wb.sheetnames:
    df_comparison = pd.read_excel(excel_file, sheet_name='比較', header=0)
    print("[比較シート]")
    print(f"  形状: {df_comparison.shape}")
    print(f"  列名: {list(df_comparison.columns)}")
    print()

    # 一致/不一致の集計
    match_counts = df_comparison['一致'].value_counts()
    print("一致/不一致の集計:")
    for match_val, count in match_counts.items():
        print(f"  {match_val}: {count}件")
    print()

    # 言語別の差異数（上位10言語）
    lang_counts = df_comparison['言語'].value_counts()
    print("言語別の差異数（上位10言語）:")
    for lang, count in lang_counts.head(10).items():
        print(f"  {lang}: {count}件")
    print()

print("=" * 80)
print("検証完了")
print("=" * 80)
