"""
各言語の翻訳をGoogle Translate APIで日本語に戻す
Excelファイルに入力と出力を両方保存（座標を一致させる）
"""
import pandas as pd
import os
from google.cloud import translate_v2 as translate
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
from dotenv import load_dotenv
import requests
import html

# .envファイルから環境変数を読み込み
load_dotenv()

# ディレクトリ設定
output_dir = r'C:\python_script\test_space\MitsubishiMultipleTranslation\output'
for_claude_dir = r'C:\python_script\test_space\MitsubishiMultipleTranslation\for_claude'

print("=" * 80)
print("Google Translate API による逆翻訳（翻訳→日本語）")
print("=" * 80)
print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print()

# 翻訳関数（APIキーを使用してREST APIで翻訳）
def translate_with_api_key(texts, source_lang, target_lang, api_key):
    """
    Google Translate REST APIを使って翻訳（複数テキスト対応）

    Args:
        texts: 翻訳するテキスト（文字列またはリスト）
        source_lang: ソース言語コード
        target_lang: ターゲット言語コード
        api_key: Google API キー

    Returns:
        翻訳結果（文字列またはリスト）
    """
    base_url = "https://translation.googleapis.com/language/translate/v2"

    # 単一文字列の場合はリストに変換
    single_input = isinstance(texts, str)
    if single_input:
        texts = [texts]

    params = {
        "key": api_key,
        "q": texts,
        "source": source_lang,
        "target": target_lang,
        "format": "text"
    }

    response = requests.post(base_url, params=params)
    response.raise_for_status()

    result = response.json()
    # HTMLエスケープをデコード
    translations = [html.unescape(t['translatedText']) for t in result['data']['translations']]

    # 単一入力の場合は文字列を返す
    return translations[0] if single_input else translations


# Google Translate クライアント初期化
api_key = os.getenv('GOOGLE_API_KEY')
use_api_key = bool(api_key)

if use_api_key:
    print("Google Translate API: 初期化成功")
    print("認証方法: APIキー（.envファイルから読み込み）")
    print("使用方式: REST API")
    translate_client = None  # REST APIを使うのでクライアント不要
else:
    try:
        # サービスアカウント（デフォルト認証）で認証
        translate_client = translate.Client()
        print("Google Translate API クライアント: 初期化成功")
        print("認証方法: サービスアカウント（.envファイルから読み込み）")
        print("使用方式: Python Client Library")
    except Exception as e:
        print(f"Google Translate API クライアント初期化エラー: {e}")
        print()
        print("注意: Google Cloud Translation API の認証が必要です。")
        print(".envファイルに以下のいずれかを設定してください：")
        print()
        print("方法1: サービスアカウントJSONファイル")
        print("  GOOGLE_APPLICATION_CREDENTIALS=C:\\path\\to\\service-account-key.json")
        print()
        print("方法2: APIキー")
        print("  GOOGLE_API_KEY=your_api_key_here")
        print()
        print(".envファイルの場所: プロジェクトルート")
        exit(1)

print()

# テンプレートCSV読み込み
input_csv = os.path.join(output_dir, '全言語統合_テンプレート_インポート用.csv')
df = pd.read_csv(input_csv, encoding='utf-8-sig')

print(f"入力CSV: {os.path.basename(input_csv)}")
print(f"行数: {len(df)}行")
print(f"列数: {len(df.columns)}列")
print()

# 言語コードと列名のマッピング
lang_mapping = {
    'en': '英語',
    'fil-PH': 'タガログ語',
    'zh': '中国語',
    'th': 'タイ語',
    'vi': 'ベトナム語',
    'my': 'ミャンマー語',
    'id': 'インドネシア語',
    'km': 'カンボジア語'
}

# Google Translate APIの言語コードマッピング（fil-PH -> tl など）
api_lang_codes = {
    'en': 'en',
    'fil-PH': 'tl',  # Tagalog
    'zh': 'zh-CN',   # Chinese (Simplified)
    'th': 'th',
    'vi': 'vi',
    'my': 'my',
    'id': 'id',
    'km': 'km'
}

# 出力用データフレーム（元のデータをコピー）
df_output = df.copy()

print("翻訳対象の列:")
for col_code, lang_name in lang_mapping.items():
    print(f"  {col_code:10s} → {lang_name}")
print()

# 各言語列を日本語に逆翻訳
print("=" * 80)
print("逆翻訳処理開始")
print("=" * 80)
print()

for col_code, lang_name in lang_mapping.items():
    print(f"処理中: {lang_name} ({col_code})")

    # 翻訳対象のテキストを取得（空欄を除く）
    translations = df[col_code].dropna().tolist()

    # 空文字列を除外
    translations = [t for t in translations if isinstance(t, str) and t.strip() != '']

    print(f"  翻訳対象: {len(translations)}件")

    if len(translations) == 0:
        print(f"  スキップ: 翻訳対象なし")
        print()
        continue

    # 逆翻訳結果を格納
    reverse_translations = {}

    # バッチ処理（API制限対策）
    # バッチサイズを20に設定（URLが長すぎるエラーを防ぐ）
    batch_size = 20
    total_batches = (len(translations) + batch_size - 1) // batch_size

    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min((batch_idx + 1) * batch_size, len(translations))
        batch = translations[start_idx:end_idx]

        print(f"  バッチ {batch_idx + 1}/{total_batches}: {len(batch)}件")

        try:
            # Google Translate API で翻訳
            source_lang = api_lang_codes[col_code]

            # 未翻訳のテキストのみを抽出
            texts_to_translate = [t for t in batch if t not in reverse_translations]

            if len(texts_to_translate) > 0:
                if use_api_key:
                    # REST APIで翻訳（APIキー使用・バッチ処理）
                    translated_texts = translate_with_api_key(
                        texts_to_translate,
                        source_lang=source_lang,
                        target_lang='ja',
                        api_key=api_key
                    )
                    # 結果を辞書に格納
                    for original, translated in zip(texts_to_translate, translated_texts):
                        reverse_translations[original] = translated
                else:
                    # Client Libraryで翻訳（サービスアカウント使用）
                    for text in texts_to_translate:
                        result = translate_client.translate(
                            text,
                            source_language=source_lang,
                            target_language='ja'
                        )
                        reverse_translations[text] = result['translatedText']

            # API制限対策（1秒待機）
            if batch_idx < total_batches - 1:
                time.sleep(1)

        except Exception as e:
            print(f"  エラー: {e}")
            print(f"  バッチ {batch_idx + 1} をスキップします")

    # 逆翻訳結果をデータフレームに適用
    df_output[col_code] = df[col_code].apply(
        lambda x: reverse_translations.get(x, x) if pd.notna(x) and x != '' else x
    )

    print(f"  完了: {len(reverse_translations)}件の逆翻訳")
    print()

print("=" * 80)
print("逆翻訳処理完了")
print("=" * 80)
print()

# Excelファイルに保存
excel_file = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')

print("Excelファイル作成中...")

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    # inputシート: 元のデータ
    df.to_excel(writer, sheet_name='input', index=False)

    # outputシート: 逆翻訳結果
    df_output.to_excel(writer, sheet_name='output', index=False)

    print(f"  inputシート: {len(df)}行 x {len(df.columns)}列")
    print(f"  outputシート: {len(df_output)}行 x {len(df_output.columns)}列")

# スタイル適用（見やすくするため）
wb = openpyxl.load_workbook(excel_file)

# inputシートのヘッダーを青色に
ws_input = wb['input']
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws_input[1]:
    cell.fill = blue_fill
    cell.font = Font(bold=True)

# outputシートのヘッダーを緑色に
ws_output = wb['output']
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
for cell in ws_output[1]:
    cell.fill = green_fill
    cell.font = Font(bold=True)

# 列幅を自動調整
for ws in [ws_input, ws_output]:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(excel_file)

print()
print(f"保存: {excel_file}")
print()

# サマリー情報をテキストファイルに保存
summary_file = os.path.join(for_claude_dir, 'reverse_translation_summary.txt')

with open(summary_file, 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("逆翻訳サマリー\n")
    f.write("=" * 80 + "\n")
    f.write(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write("\n")

    f.write(f"入力CSV: {os.path.basename(input_csv)}\n")
    f.write(f"総行数: {len(df)}行\n")
    f.write("\n")

    f.write("逆翻訳した言語:\n")
    for col_code, lang_name in lang_mapping.items():
        count = df[col_code].notna().sum()
        f.write(f"  {lang_name:15s} ({col_code:10s}): {count}件\n")
    f.write("\n")

    f.write(f"出力Excel: {os.path.basename(excel_file)}\n")
    f.write("  - inputシート: 元のCSVデータ\n")
    f.write("  - outputシート: 逆翻訳結果（各言語→日本語）\n")
    f.write("\n")

    f.write("座標: 両シートの行・列位置は完全に一致\n")

print(f"サマリー保存: {summary_file}")
print()

print("=" * 80)
print("完了")
print("=" * 80)
