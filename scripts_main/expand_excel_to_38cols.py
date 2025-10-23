"""
Excelの翻訳・再翻訳シートを38列に拡張
比較シートは空セルを無視
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

project_root = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
core_files_dir = os.path.join(project_root, 'core_files')
output_dir = os.path.join(project_root, 'output')
deliverables_dir = os.path.join(project_root, '成果物_20251023')

print("=" * 80)
print("Excelシートを38列に拡張 + 比較シート作成")
print("=" * 80)
print()

# ========================================
# Step 1: 38列のヘッダーを取得
# ========================================
print("[1/4] 38列のヘッダーを取得")
print("-" * 80)

source_csv = os.path.join(core_files_dir, 'output_csv_template_utf8bom.csv')
df_template = pd.read_csv(source_csv, encoding='utf-8-sig')

# 翻訳言語数列を削除
if '翻訳言語数' in df_template.columns:
    column_names_38 = [col for col in df_template.columns if col != '翻訳言語数']
else:
    column_names_38 = df_template.columns.tolist()

print(f"38列のヘッダー: {len(column_names_38)}列")
print(f"  先頭5列: {column_names_38[:5]}")
print(f"  末尾5列: {column_names_38[-5:]}")
print()

# ========================================
# Step 2: 既存の逆翻訳Excelを読み込み
# ========================================
print("[2/4] 既存の逆翻訳Excelを読み込み")
print("-" * 80)

excel_file = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')

# 翻訳シートと再翻訳シートを読み込み
df_translation = pd.read_excel(excel_file, sheet_name='翻訳', header=0)
df_retranslation = pd.read_excel(excel_file, sheet_name='再翻訳', header=0)

print(f"翻訳シート: {len(df_translation)}行 x {len(df_translation.columns)}列")
print(f"再翻訳シート: {len(df_retranslation)}行 x {len(df_retranslation.columns)}列")
print()

# ========================================
# Step 3: 38列に拡張
# ========================================
print("[3/4] 38列に拡張")
print("-" * 80)

# 新しいDataFrameを作成（38列）
df_translation_38 = pd.DataFrame(index=df_translation.index, columns=column_names_38)
df_retranslation_38 = pd.DataFrame(index=df_retranslation.index, columns=column_names_38)

# 既存のデータをコピー
for col in df_translation.columns:
    if col in column_names_38:
        df_translation_38[col] = df_translation[col]
        df_retranslation_38[col] = df_retranslation[col]

print(f"翻訳シート（拡張後）: {len(df_translation_38)}行 x {len(df_translation_38.columns)}列")
print(f"再翻訳シート（拡張後）: {len(df_retranslation_38)}行 x {len(df_retranslation_38.columns)}列")
print()

# ========================================
# Step 4: 比較データを作成（空セルを無視）
# ========================================
print("[4/4] 比較シート作成（空セルを無視）")
print("-" * 80)

comparison_data = []

# 各行を比較
for row_idx in range(len(df_translation_38)):
    japanese_word = df_translation_38.loc[row_idx, 'ja']

    # 各言語列を比較（日本語以外）
    for col_idx, lang in enumerate(column_names_38):
        if lang == 'ja':
            continue

        # 翻訳と再翻訳の値を取得
        translation_val = df_translation_38.loc[row_idx, lang]
        retranslation_val = df_retranslation_38.loc[row_idx, lang]

        # 空セルをスキップ
        if pd.isna(translation_val) or str(translation_val).strip() == '':
            continue
        if pd.isna(retranslation_val) or str(retranslation_val).strip() == '':
            continue

        # 文字列に変換
        translation_str = str(translation_val).strip()
        retranslation_str = str(retranslation_val).strip()

        # 差異がある場合のみ記録
        if translation_str != retranslation_str:
            # セルアドレスを計算
            col_letter = chr(65 + column_names_38.index(lang))  # A=ja, B=en, ...
            cell_address = f"{col_letter}{row_idx + 2}"  # +2 (ヘッダー1行 + 0-indexed)

            # 単語と再翻訳の一致判定
            match_result = japanese_word.strip() == retranslation_str.strip()

            comparison_data.append({
                'アドレス': cell_address,
                '言語': lang,
                '単語': japanese_word,
                '再翻訳': retranslation_str,
                '翻訳': translation_str,
                '一致': match_result
            })

print(f"差異検出: {len(comparison_data)}件（空セルを除外）")
print()

# 比較DataFrameを作成
df_comparison = pd.DataFrame(comparison_data)

# ========================================
# Excelに出力
# ========================================
print("Excelファイル作成中...")
print()

output_excel = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

# Excelファイルを作成
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    # 翻訳シート（38列、ヘッダーあり）
    df_translation_38.to_excel(writer, sheet_name='翻訳', index=False, header=True)
    # 再翻訳シート（38列、ヘッダーあり）
    df_retranslation_38.to_excel(writer, sheet_name='再翻訳', index=False, header=True)

print(f"  翻訳シート: {len(df_translation_38)+1}行（ヘッダー1行 + データ{len(df_translation_38)}行）")
print(f"  再翻訳シート: {len(df_retranslation_38)+1}行（ヘッダー1行 + データ{len(df_retranslation_38)}行）")
print()

# スタイル適用
print("スタイル適用中...")
wb = openpyxl.load_workbook(output_excel)

# 翻訳シートのヘッダーを青色に
ws_translation = wb['翻訳']
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws_translation[1]:
    cell.fill = blue_fill
    cell.font = Font(bold=True)

# 再翻訳シートのヘッダーを緑色に
ws_retranslation = wb['再翻訳']
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
for cell in ws_retranslation[1]:
    cell.fill = green_fill
    cell.font = Font(bold=True)

# 列幅を自動調整
for ws in [ws_translation, ws_retranslation]:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# 比較シートを追加
if len(df_comparison) > 0:
    print("比較シート追加中...")

    # 比較シートを作成（最後に追加）
    ws_comparison = wb.create_sheet('比較')

    # ヘッダー設定
    headers = ['アドレス', '言語', '単語', '再翻訳', '翻訳', '一致']
    ws_comparison.append(headers)

    # ヘッダーのスタイル設定（黄色背景）
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws_comparison[1]:
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # データを追加
    for idx, row in df_comparison.iterrows():
        ws_comparison.append([
            row['アドレス'],
            row['言語'],
            row['単語'],
            row['再翻訳'],
            row['翻訳'],
            row['一致']
        ])

    # 列幅を調整
    ws_comparison.column_dimensions['A'].width = 10  # アドレス
    ws_comparison.column_dimensions['B'].width = 10  # 言語
    ws_comparison.column_dimensions['C'].width = 30  # 単語
    ws_comparison.column_dimensions['D'].width = 40  # 再翻訳
    ws_comparison.column_dimensions['E'].width = 40  # 翻訳
    ws_comparison.column_dimensions['F'].width = 10  # 一致

# 保存
wb.save(output_excel)
print()
print(f"保存: {output_excel}")

# outputフォルダにもコピー
output_excel_copy = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')
wb.save(output_excel_copy)
print(f"コピー保存: {output_excel_copy}")
print()

print("=" * 80)
print("完了")
print("=" * 80)
print()

print("シート構成:")
print(f"  1. 翻訳（青ヘッダー）: {len(df_translation_38)}行 x {len(df_translation_38.columns)}列")
print(f"  2. 再翻訳（緑ヘッダー）: {len(df_retranslation_38)}行 x {len(df_retranslation_38.columns)}列")
if len(df_comparison) > 0:
    print(f"  3. 比較（黄ヘッダー）: {len(df_comparison)}件の差異（空セル除外）")
print()
