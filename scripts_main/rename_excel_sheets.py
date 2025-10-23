"""
Excelシート名変更: input → 翻訳, output → 再翻訳
"""
import openpyxl
import os

project_root = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
deliverables_dir = os.path.join(project_root, '成果物_20251023')
output_dir = os.path.join(project_root, 'output')

print("=" * 80)
print("Excelシート名変更")
print("=" * 80)
print()

# 成果物フォルダのExcel
excel_file = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)

    # シート名変更
    if 'input' in wb.sheetnames:
        wb['input'].title = '翻訳'
        print("OK: 'input' -> '翻訳'")

    if 'output' in wb.sheetnames:
        wb['output'].title = '再翻訳'
        print("OK: 'output' -> '再翻訳'")

    wb.save(excel_file)
    print()
    print(f"保存: {excel_file}")
else:
    print(f"エラー: ファイルが見つかりません: {excel_file}")

print()

# outputフォルダのExcelもコピー
excel_file_output = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')

if os.path.exists(excel_file_output):
    wb = openpyxl.load_workbook(excel_file_output)

    # シート名変更
    if 'input' in wb.sheetnames:
        wb['input'].title = '翻訳'

    if 'output' in wb.sheetnames:
        wb['output'].title = '再翻訳'

    wb.save(excel_file_output)
    print(f"コピー保存: {excel_file_output}")

print()
print("=" * 80)
print("シート名変更完了")
print("=" * 80)
print()
print("新しいシート名:")
print("  - 翻訳（元: input）: 元の翻訳データ")
print("  - 再翻訳（元: output）: 日本語逆翻訳データ")
print()
