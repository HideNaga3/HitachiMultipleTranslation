"""
ヘッダー付きのExcelファイルを作成
"""
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

output_dir = r'C:\python_script\test_space\MitsubishiMultipleTranslation\output'
deliverables_dir = r'C:\python_script\test_space\MitsubishiMultipleTranslation\成果物_20251023'

print("=" * 80)
print("ヘッダー付きExcelファイルの作成")
print("=" * 80)
print()

# 逆翻訳結果読み込み（元のファイルはヘッダーあり）
excel_file = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')

# 元ファイルを読み込む前に、最初に作成したバージョンを確認
# reverse_translate_to_japanese.py が作成した元のファイル

# テンプレートCSVから列名を取得
input_csv = os.path.join(output_dir, '全言語統合_テンプレート_インポート用.csv')
df_template = pd.read_csv(input_csv, encoding='utf-8-sig')
column_names = df_template.columns.tolist()

print(f"列名（{len(column_names)}列）:")
for i, col in enumerate(column_names):
    print(f"  {i+1}. {col}")
print()

# データ読み込み（ヘッダーなしで読み込んで、後で列名を設定）
df_input = pd.read_excel(excel_file, sheet_name='input', header=None)
df_output = pd.read_excel(excel_file, sheet_name='output', header=None)

# 列名を設定
df_input.columns = column_names
df_output.columns = column_names

print(f"inputシート: {len(df_input)}行 x {len(df_input.columns)}列")
print(f"outputシート: {len(df_output)}行 x {len(df_output.columns)}列")
print()

# ヘッダー付きExcelファイル作成
output_excel = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

print("ヘッダー付きExcelファイル作成中...")

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    # inputシート（ヘッダーあり）
    df_input.to_excel(writer, sheet_name='input', index=False, header=True)
    # outputシート（ヘッダーあり）
    df_output.to_excel(writer, sheet_name='output', index=False, header=True)

print(f"  inputシート: {len(df_input)+1}行（ヘッダー1行 + データ{len(df_input)}行）")
print(f"  outputシート: {len(df_output)+1}行（ヘッダー1行 + データ{len(df_output)}行）")
print()

# スタイル適用
wb = openpyxl.load_workbook(output_excel)

# inputシートのヘッダーを青色に
ws_input = wb['input']
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws_input[1]:
    cell.fill = blue_fill
    cell.font = Font(bold=True)

# outputシートのヘッダーを緑色に
ws_output = wb['output']
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
for cell in ws_output[1]:
    cell.fill = green_fill
    cell.font = Font(bold=True)

# 列幅を自動調整
for ws in [ws_input, ws_output]:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_excel)

print(f"保存: {output_excel}")
print()

# outputにもコピー
output_excel_copy = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')
wb.save(output_excel_copy)
print(f"コピー保存: {output_excel_copy}")
print()

print("=" * 80)
print("完了")
print("=" * 80)
print()
print("注意:")
print("  - inputシート: ヘッダーあり（青色）+ データ524行")
print("  - outputシート: ヘッダーあり（緑色）+ データ524行")
print("  - ヘッダー: ja, en, fil-PH, zh, th, vi, my, id, km")
print()
