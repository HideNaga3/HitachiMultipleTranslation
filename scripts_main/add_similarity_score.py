"""
比較シートに類似度スコアを追加
- 文字列類似度（difflib）: 0-1の範囲
- 意味的に近いかどうかの判定
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from difflib import SequenceMatcher

project_root = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
deliverables_dir = os.path.join(project_root, '成果物_20251023')
output_dir = os.path.join(project_root, 'output')

print("=" * 80)
print("比較シートに類似度スコアを追加")
print("=" * 80)
print()

# ========================================
# 類似度計算関数
# ========================================

def calculate_similarity(str1, str2):
    """
    2つの文字列の類似度を計算（0-1の範囲）

    Parameters:
    -----------
    str1, str2 : str
        比較する文字列

    Returns:
    --------
    float : 類似度スコア（0-1）
    """
    if not str1 or not str2:
        return 0.0

    # SequenceMatcherで類似度を計算
    return SequenceMatcher(None, str1, str2).ratio()


def classify_similarity(similarity_score, exact_match):
    """
    類似度スコアから分類

    Parameters:
    -----------
    similarity_score : float
        類似度スコア（0-1）
    exact_match : bool
        完全一致かどうか

    Returns:
    --------
    str : 分類（完全一致、高類似、中類似、低類似、不一致）
    """
    if exact_match:
        return "完全一致"
    elif similarity_score >= 0.8:
        return "高類似"
    elif similarity_score >= 0.6:
        return "中類似"
    elif similarity_score >= 0.4:
        return "低類似"
    else:
        return "不一致"


# ========================================
# 比較シート読み込み
# ========================================
print("比較シートを読み込み中...")

excel_file = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')
df_comparison = pd.read_excel(excel_file, sheet_name='比較', header=0)

print(f"比較シート: {len(df_comparison)}行")
print()

# ========================================
# 類似度計算
# ========================================
print("類似度を計算中...")
print()

# 類似度スコアを計算
similarity_scores = []
similarity_classes = []

for idx, row in df_comparison.iterrows():
    word = str(row['単語']).strip()
    retranslation = str(row['再翻訳']).strip()
    exact_match = row['一致']

    # 類似度計算
    score = calculate_similarity(word, retranslation)
    similarity_scores.append(score)

    # 分類
    classification = classify_similarity(score, exact_match)
    similarity_classes.append(classification)

# DataFrameに追加（パーセンテージ表記）
df_comparison['類似度_difflib'] = [f"{score*100:.1f}%" for score in similarity_scores]
df_comparison['分類'] = similarity_classes
df_comparison['_similarity_raw'] = similarity_scores  # 一時的に計算用に保持

print(f"類似度計算完了: {len(df_comparison)}件")
print()

# ========================================
# 統計情報
# ========================================
print("=" * 80)
print("統計情報")
print("=" * 80)
print()

# 分類の集計
class_counts = df_comparison['分類'].value_counts()
print("分類別の件数:")
for classification, count in class_counts.items():
    print(f"  {classification}: {count}件")

print()

# 類似度の統計（生の値を使用）
print("類似度の統計:")
print(f"  平均: {df_comparison['_similarity_raw'].mean()*100:.1f}%")
print(f"  中央値: {df_comparison['_similarity_raw'].median()*100:.1f}%")
print(f"  最小: {df_comparison['_similarity_raw'].min()*100:.1f}%")
print(f"  最大: {df_comparison['_similarity_raw'].max()*100:.1f}%")
print()

# 高類似のサンプルを表示
print("高類似（0.8以上）のサンプル（最初の10件）:")
print("-" * 80)
high_similarity = df_comparison[df_comparison['_similarity_raw'] >= 0.8].head(10)
for idx, row in high_similarity.iterrows():
    print(f"[{row['アドレス']}] {row['言語']} | 類似度: {row['類似度_difflib']}")
    print(f"  単語: {row['単語']}")
    print(f"  再翻訳: {row['再翻訳']}")
    print()

print()

# ========================================
# Excelファイルを更新
# ========================================
print("Excelファイルを更新中...")

# 既存のExcelファイルを開く
wb = openpyxl.load_workbook(excel_file)

# 比較シートを削除
if '比較' in wb.sheetnames:
    wb.remove(wb['比較'])

# 新しい比較シートを作成（最後に追加）
ws_comparison = wb.create_sheet('比較')

# ヘッダー設定
headers = ['アドレス', '言語', '単語', '再翻訳', '翻訳', '一致', '類似度_difflib', '分類']
ws_comparison.append(headers)

# ヘッダーのスタイル設定（黄色背景）
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for cell in ws_comparison[1]:
    cell.fill = yellow_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# データを追加
for idx, row in df_comparison.iterrows():
    ws_comparison.append([
        row['アドレス'],
        row['言語'],
        row['単語'],
        row['再翻訳'],
        row['翻訳'],
        row['一致'],
        row['類似度_difflib'],
        row['分類']
    ])

# 列幅を調整
ws_comparison.column_dimensions['A'].width = 10  # アドレス
ws_comparison.column_dimensions['B'].width = 10  # 言語
ws_comparison.column_dimensions['C'].width = 30  # 単語
ws_comparison.column_dimensions['D'].width = 40  # 再翻訳
ws_comparison.column_dimensions['E'].width = 40  # 翻訳
ws_comparison.column_dimensions['F'].width = 10  # 一致
ws_comparison.column_dimensions['G'].width = 15  # 類似度_difflib
ws_comparison.column_dimensions['H'].width = 12  # 分類

# 保存
wb.save(excel_file)
print(f"保存: {excel_file}")

# outputフォルダにもコピー
output_excel_copy = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')
wb.save(output_excel_copy)
print(f"コピー保存: {output_excel_copy}")
print()

print("=" * 80)
print("完了")
print("=" * 80)
print()

print("比較シート構成:")
print("  1. アドレス - セル位置")
print("  2. 言語 - 言語コード")
print("  3. 単語 - 元の日本語")
print("  4. 再翻訳 - 逆翻訳された日本語")
print("  5. 翻訳 - 元の翻訳")
print("  6. 一致 - 完全一致かどうか（TRUE/FALSE）")
print("  7. 類似度_difflib - 文字列類似度スコア（パーセンテージ: xx.x%）")
print("  8. 分類 - 完全一致、高類似、中類似、低類似、不一致")
print()

print("分類基準:")
print("  - 完全一致: 一致 = TRUE")
print("  - 高類似: 類似度 >= 80.0%")
print("  - 中類似: 類似度 >= 60.0%")
print("  - 低類似: 類似度 >= 40.0%")
print("  - 不一致: 類似度 < 40.0%")
print()
