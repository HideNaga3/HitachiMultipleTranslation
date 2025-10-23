"""
成果物出力関数
- create_import_csv_only(): インポート用CSVのみ出力
- create_import_csv_with_comparison_excel(): インポート用CSV + 比較Excel出力
"""
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# パス設定
PROJECT_ROOT = r'C:\python_script\test_space\MitsubishiMultipleTranslation'
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'output')
CORE_FILES_DIR = os.path.join(PROJECT_ROOT, 'core_files')
DELIVERABLES_DIR = os.path.join(PROJECT_ROOT, '成果物_20251023')


def create_import_csv_only(output_path=None):
    """
    インポート用CSV（38列）のみを出力

    Parameters:
    -----------
    output_path : str, optional
        出力先パス。指定しない場合は成果物フォルダに出力

    Returns:
    --------
    str : 出力ファイルパス
    """
    print("=" * 80)
    print("インポート用CSV作成（38列）")
    print("=" * 80)
    print()

    # core_filesのCSVを読み込み
    source_file = os.path.join(CORE_FILES_DIR, 'output_csv_template_utf8bom.csv')

    if not os.path.exists(source_file):
        raise FileNotFoundError(f"ソースファイルが見つかりません: {source_file}")

    df = pd.read_csv(source_file, encoding='utf-8-sig')

    print(f"ソースファイル: {os.path.basename(source_file)}")
    print(f"  行数: {len(df)}行")
    print(f"  列数: {len(df.columns)}列")
    print()

    # 翻訳言語数列を削除
    if '翻訳言語数' in df.columns:
        df_import = df.drop(columns=['翻訳言語数'])
        print("OK: '翻訳言語数'列を削除")
    else:
        df_import = df.copy()
        print("警告: '翻訳言語数'列が見つかりません")

    print()
    print(f"インポート用CSV:")
    print(f"  行数: {len(df_import)}行")
    print(f"  列数: {len(df_import.columns)}列")
    print()

    # 出力先決定
    if output_path is None:
        output_path = os.path.join(DELIVERABLES_DIR, '01_全言語統合_テンプレート_インポート用.csv')

    # 保存
    df_import.to_csv(output_path, index=False, encoding='utf-8-sig')

    file_size_kb = os.path.getsize(output_path) / 1024
    print(f"出力: {output_path}")
    print(f"ファイルサイズ: {file_size_kb:.2f} KB")
    print()

    print("=" * 80)
    print("CSV作成完了")
    print("=" * 80)
    print()

    return output_path


def create_import_csv_with_comparison_excel(csv_output_path=None, excel_output_path=None):
    """
    インポート用CSV + 比較Excel（シート名: 翻訳, 再翻訳）を出力

    Parameters:
    -----------
    csv_output_path : str, optional
        CSV出力先パス。指定しない場合は成果物フォルダに出力
    excel_output_path : str, optional
        Excel出力先パス。指定しない場合は成果物フォルダに出力

    Returns:
    --------
    tuple : (csv_path, excel_path)
    """
    print("=" * 80)
    print("インポート用CSV + 比較Excel作成")
    print("=" * 80)
    print()

    # ========================================
    # Part 1: インポート用CSV作成
    # ========================================
    print("[1/2] インポート用CSV作成")
    print("-" * 80)
    csv_path = create_import_csv_only(csv_output_path)
    print()

    # ========================================
    # Part 2: 比較Excel作成
    # ========================================
    print("[2/2] 比較Excel作成")
    print("-" * 80)
    print()

    # インポート用CSVから列名を取得
    df_template = pd.read_csv(csv_path, encoding='utf-8-sig')
    column_names = df_template.columns.tolist()

    # 9列のみ抽出（ja + 8言語）
    template_columns = ['ja', 'en', 'fil-PH', 'zh', 'th', 'vi', 'my', 'id', 'km']

    # 元の逆翻訳Excelを読み込み
    source_excel = os.path.join(OUTPUT_DIR, '逆翻訳_検証結果.xlsx')

    if not os.path.exists(source_excel):
        raise FileNotFoundError(f"ソースExcelが見つかりません: {source_excel}")

    # データ読み込み（ヘッダーあり）
    # シート名は '翻訳' と '再翻訳' に変更済み
    try:
        df_input = pd.read_excel(source_excel, sheet_name='翻訳', header=0)
        df_output = pd.read_excel(source_excel, sheet_name='再翻訳', header=0)
    except ValueError:
        # 旧シート名で試行
        df_input = pd.read_excel(source_excel, sheet_name='input', header=0)
        df_output = pd.read_excel(source_excel, sheet_name='output', header=0)

    print(f"ソースExcel: {os.path.basename(source_excel)}")
    print(f"  翻訳シート: {len(df_input)}行 x {len(df_input.columns)}列")
    print(f"  再翻訳シート: {len(df_output)}行 x {len(df_output.columns)}列")
    print()

    # 出力先決定
    if excel_output_path is None:
        excel_output_path = os.path.join(DELIVERABLES_DIR, '02_逆翻訳_検証結果.xlsx')

    # Excel作成
    print("Excel作成中...")
    with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
        # 翻訳シート（ヘッダーあり）
        df_input.to_excel(writer, sheet_name='翻訳', index=False, header=True)
        # 再翻訳シート（ヘッダーあり）
        df_output.to_excel(writer, sheet_name='再翻訳', index=False, header=True)

    print(f"  翻訳シート: {len(df_input)+1}行（ヘッダー1行 + データ{len(df_input)}行）")
    print(f"  再翻訳シート: {len(df_output)+1}行（ヘッダー1行 + データ{len(df_output)}行）")
    print()

    # スタイル適用
    print("スタイル適用中...")
    wb = openpyxl.load_workbook(excel_output_path)

    # 翻訳シートのヘッダーを青色に
    ws_input = wb['翻訳']
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    for cell in ws_input[1]:
        cell.fill = blue_fill
        cell.font = Font(bold=True)

    # 再翻訳シートのヘッダーを緑色に
    ws_output = wb['再翻訳']
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    for cell in ws_output[1]:
        cell.fill = green_fill
        cell.font = Font(bold=True)

    # 列幅を自動調整
    for ws in [ws_input, ws_output]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_output_path)
    print()

    file_size_kb = os.path.getsize(excel_output_path) / 1024
    print(f"出力: {excel_output_path}")
    print(f"ファイルサイズ: {file_size_kb:.2f} KB")
    print()

    # outputフォルダにもコピー
    excel_copy_path = os.path.join(OUTPUT_DIR, '逆翻訳_検証結果.xlsx')
    wb.save(excel_copy_path)
    print(f"コピー保存: {excel_copy_path}")
    print()

    print("=" * 80)
    print("CSV + Excel作成完了")
    print("=" * 80)
    print()

    return csv_path, excel_output_path


# スタンドアローン実行時
if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        mode = sys.argv[1]

        if mode == "csv":
            # CSVのみ出力
            create_import_csv_only()
        elif mode == "both":
            # CSV + Excel出力
            create_import_csv_with_comparison_excel()
        else:
            print("使用方法:")
            print("  python output_functions.py csv     # CSVのみ出力")
            print("  python output_functions.py both    # CSV + Excel出力")
    else:
        # デフォルト: 両方出力
        print("デフォルトモード: CSV + Excel出力")
        print()
        create_import_csv_with_comparison_excel()
