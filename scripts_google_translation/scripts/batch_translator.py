"""
CSV/Excelファイルからの一括翻訳スクリプト
往復翻訳と完全一致判定に対応
複数ファイルの一括処理に対応
"""
import csv
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Union
from translator import round_trip_translate_batch


def check_csv_structure(file_paths: List[str], encoding: str = "utf-8-sig") -> Dict:
    """
    複数のCSVファイルの列構造が一致しているかチェック

    Args:
        file_paths (List[str]): チェックするCSVファイルパスのリスト
        encoding (str): ファイルのエンコーディング（デフォルト: "utf-8-sig"）

    Returns:
        dict: チェック結果
            {
                "is_valid": True/False（全ファイルの構造が一致しているか）,
                "column_names": 列名のリスト（最初のファイルの列名）,
                "column_count": 列数,
                "files": [
                    {
                        "path": ファイルパス,
                        "column_names": 列名リスト,
                        "column_count": 列数,
                        "is_match": True/False
                    },
                    ...
                ],
                "error_files": 不一致のファイルパスのリスト
            }

    Raises:
        FileNotFoundError: ファイルが見つからない場合
    """
    if not file_paths:
        raise ValueError("ファイルパスのリストが空です")

    results = {
        "is_valid": True,
        "column_names": None,
        "column_count": 0,
        "files": [],
        "error_files": []
    }

    # 最初のファイルを基準とする
    base_file = Path(file_paths[0])
    if not base_file.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_paths[0]}")

    with open(base_file, 'r', encoding=encoding, newline='') as f:
        reader = csv.reader(f)
        try:
            base_columns = next(reader)
            results["column_names"] = base_columns
            results["column_count"] = len(base_columns)
        except StopIteration:
            raise ValueError(f"ファイルが空です: {file_paths[0]}")

    # 全ファイルをチェック
    for file_path in file_paths:
        file = Path(file_path)
        if not file.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        with open(file, 'r', encoding=encoding, newline='') as f:
            reader = csv.reader(f)
            try:
                columns = next(reader)
                is_match = (columns == results["column_names"])

                file_info = {
                    "path": str(file_path),
                    "column_names": columns,
                    "column_count": len(columns),
                    "is_match": is_match
                }
                results["files"].append(file_info)

                if not is_match:
                    results["is_valid"] = False
                    results["error_files"].append(str(file_path))
            except StopIteration:
                results["is_valid"] = False
                results["error_files"].append(str(file_path))
                results["files"].append({
                    "path": str(file_path),
                    "column_names": [],
                    "column_count": 0,
                    "is_match": False
                })

    return results


def check_excel_structure(file_paths: List[str], sheet_name: Optional[str] = None) -> Dict:
    """
    複数のExcelファイルの列構造が一致しているかチェック

    Args:
        file_paths (List[str]): チェックするExcelファイルパスのリスト
        sheet_name (str): チェックするシート名（省略時は最初のシート）

    Returns:
        dict: チェック結果（check_csv_structureと同じ形式）

    Raises:
        FileNotFoundError: ファイルが見つからない場合
        ImportError: openpyxlがインストールされていない場合
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxlがインストールされていません。'pip install openpyxl'を実行してください。")

    if not file_paths:
        raise ValueError("ファイルパスのリストが空です")

    results = {
        "is_valid": True,
        "column_names": None,
        "column_count": 0,
        "files": [],
        "error_files": []
    }

    # 最初のファイルを基準とする
    base_file = Path(file_paths[0])
    if not base_file.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_paths[0]}")

    wb = load_workbook(base_file, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が見つかりません。利用可能なシート: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # ヘッダー行を取得
    base_columns = None
    for row in ws.iter_rows(values_only=True, max_row=1):
        base_columns = [str(cell) if cell is not None else "" for cell in row]
        break
    wb.close()

    if base_columns is None:
        raise ValueError(f"ファイルが空です: {file_paths[0]}")

    results["column_names"] = base_columns
    results["column_count"] = len(base_columns)

    # 全ファイルをチェック
    for file_path in file_paths:
        file = Path(file_path)
        if not file.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        wb = load_workbook(file, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                results["is_valid"] = False
                results["error_files"].append(str(file_path))
                results["files"].append({
                    "path": str(file_path),
                    "column_names": [],
                    "column_count": 0,
                    "is_match": False
                })
                wb.close()
                continue
            ws = wb[sheet_name]
        else:
            ws = wb.active

        columns = None
        for row in ws.iter_rows(values_only=True, max_row=1):
            columns = [str(cell) if cell is not None else "" for cell in row]
            break
        wb.close()

        if columns is None:
            results["is_valid"] = False
            results["error_files"].append(str(file_path))
            results["files"].append({
                "path": str(file_path),
                "column_names": [],
                "column_count": 0,
                "is_match": False
            })
        else:
            is_match = (columns == results["column_names"])
            file_info = {
                "path": str(file_path),
                "column_names": columns,
                "column_count": len(columns),
                "is_match": is_match
            }
            results["files"].append(file_info)

            if not is_match:
                results["is_valid"] = False
                results["error_files"].append(str(file_path))

    return results


def translate_from_csv(
    input_file: str,
    output_file: Optional[str] = None,
    column_index: int = 0,
    intermediate_lang: str = "en",
    encoding: str = "utf-8-sig"
) -> dict:
    """
    CSVファイルから単語を読み込んで往復翻訳を実行

    Args:
        input_file (str): 入力CSVファイルのパス
        output_file (str): 出力CSVファイルのパス（省略時は自動生成）
        column_index (int): 翻訳する列のインデックス（0始まり、デフォルト: 0）
        intermediate_lang (str): 中間言語コード（デフォルト: "en"）
        encoding (str): 入力ファイルのエンコーディング（デフォルト: "utf-8-sig"）

    Returns:
        dict: 処理結果の情報
            {
                "input_file": 入力ファイルパス,
                "output_file": 出力ファイルパス,
                "total_count": 処理件数,
                "perfect_match_count": 完全一致件数,
                "perfect_match_rate": 完全一致率
            }

    Raises:
        Exception: ファイル読み込みエラー、翻訳エラー
    """
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_file}")

    # 出力ファイル名を自動生成
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = input_path.parent / "output"
        output_dir.mkdir(exist_ok=True)
        output_file = output_dir / f"{input_path.stem}_translated_{timestamp}.csv"
    else:
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    # CSVファイルから単語を読み込み
    words = []
    with open(input_path, 'r', encoding=encoding, newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) > column_index and row[column_index].strip():
                words.append(row[column_index].strip())

    if not words:
        raise ValueError(f"列インデックス {column_index} にデータが見つかりません")

    print(f"[INFO] 読み込み完了: {len(words)}件")
    print(f"[INFO] 翻訳実行中: 日本語 → {intermediate_lang} → 日本語")

    # 往復翻訳を実行
    results = round_trip_translate_batch(words, intermediate_lang)

    # 統計情報を計算
    perfect_match_count = sum(1 for r in results if r["is_perfect_match"])
    total_count = len(results)
    perfect_match_rate = (perfect_match_count / total_count * 100) if total_count > 0 else 0

    # CSV出力
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        fieldnames = [
            "元の日本語",
            "中間言語",
            "中間言語の翻訳",
            "日本語への逆翻訳",
            "完全一致"
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for result in results:
            writer.writerow({
                "元の日本語": result["original"],
                "中間言語": result["intermediate_lang"],
                "中間言語の翻訳": result["intermediate_text"],
                "日本語への逆翻訳": result["back_translation"],
                "完全一致": result["is_perfect_match"]
            })

    return {
        "input_file": str(input_path),
        "output_file": str(output_file),
        "total_count": total_count,
        "perfect_match_count": perfect_match_count,
        "perfect_match_rate": perfect_match_rate
    }


def translate_from_excel(
    input_file: str,
    output_file: Optional[str] = None,
    column_index: int = 0,
    intermediate_lang: str = "en",
    sheet_name: Optional[str] = None
) -> dict:
    """
    Excelファイルから単語を読み込んで往復翻訳を実行

    Args:
        input_file (str): 入力Excelファイルのパス
        output_file (str): 出力CSVファイルのパス（省略時は自動生成）
        column_index (int): 翻訳する列のインデックス（0始まり、デフォルト: 0）
        intermediate_lang (str): 中間言語コード（デフォルト: "en"）
        sheet_name (str): シート名（省略時は最初のシート）

    Returns:
        dict: 処理結果の情報

    Raises:
        Exception: ファイル読み込みエラー、翻訳エラー
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxlがインストールされていません。'pip install openpyxl'を実行してください。")

    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_file}")

    # 出力ファイル名を自動生成
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = input_path.parent / "output"
        output_dir.mkdir(exist_ok=True)
        output_file = output_dir / f"{input_path.stem}_translated_{timestamp}.csv"
    else:
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    # Excelファイルから単語を読み込み
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が見つかりません。利用可能なシート: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    words = []
    for row in ws.iter_rows(values_only=True):
        if row and len(row) > column_index and row[column_index]:
            word = str(row[column_index]).strip()
            if word:
                words.append(word)

    wb.close()

    if not words:
        raise ValueError(f"列インデックス {column_index} にデータが見つかりません")

    print(f"[INFO] 読み込み完了: {len(words)}件")
    print(f"[INFO] 翻訳実行中: 日本語 → {intermediate_lang} → 日本語")

    # 往復翻訳を実行
    results = round_trip_translate_batch(words, intermediate_lang)

    # 統計情報を計算
    perfect_match_count = sum(1 for r in results if r["is_perfect_match"])
    total_count = len(results)
    perfect_match_rate = (perfect_match_count / total_count * 100) if total_count > 0 else 0

    # CSV出力
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        fieldnames = [
            "元の日本語",
            "中間言語",
            "中間言語の翻訳",
            "日本語への逆翻訳",
            "完全一致"
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for result in results:
            writer.writerow({
                "元の日本語": result["original"],
                "中間言語": result["intermediate_lang"],
                "中間言語の翻訳": result["intermediate_text"],
                "日本語への逆翻訳": result["back_translation"],
                "完全一致": result["is_perfect_match"]
            })

    return {
        "input_file": str(input_path),
        "output_file": str(output_file),
        "total_count": total_count,
        "perfect_match_count": perfect_match_count,
        "perfect_match_rate": perfect_match_rate
    }


def translate_from_multiple_csv(
    file_paths: List[str],
    output_file: Optional[str] = None,
    column_index: int = 0,
    intermediate_lang: str = "en",
    encoding: str = "utf-8-sig",
    check_structure: bool = True
) -> Dict:
    """
    複数のCSVファイルを一括で翻訳し、1つのCSVファイルにまとめて出力

    Args:
        file_paths (List[str]): 入力CSVファイルパスのリスト
        output_file (str): 出力CSVファイルのパス（省略時は自動生成）
        column_index (int): 翻訳する列のインデックス（0始まり、デフォルト: 0）
        intermediate_lang (str): 中間言語コード（デフォルト: "en"）
        encoding (str): 入力ファイルのエンコーディング（デフォルト: "utf-8-sig"）
        check_structure (bool): 列構造チェックを実行するか（デフォルト: True）

    Returns:
        dict: 処理結果
            {
                "structure_check": 構造チェック結果（check_structure=Falseの場合はNone）,
                "output_file": 出力ファイルパス,
                "total_files": 処理ファイル数,
                "total_count": 総翻訳件数,
                "perfect_match_count": 完全一致件数,
                "perfect_match_rate": 完全一致率,
                "success_count": 成功ファイル数,
                "error_count": エラーファイル数,
                "errors": エラー情報のリスト
            }

    Raises:
        ValueError: ファイルリストが空、または構造チェックで不一致が検出された場合
    """
    if not file_paths:
        raise ValueError("ファイルパスのリストが空です")

    result_data = {
        "structure_check": None,
        "output_file": None,
        "total_files": len(file_paths),
        "total_count": 0,
        "perfect_match_count": 0,
        "perfect_match_rate": 0.0,
        "success_count": 0,
        "error_count": 0,
        "errors": []
    }

    # 列構造チェック
    if check_structure:
        print("[INFO] 列構造チェック実行中...")
        structure_check = check_csv_structure(file_paths, encoding)
        result_data["structure_check"] = structure_check

        if not structure_check["is_valid"]:
            print("[ERROR] 列構造が一致していません")
            print(f"  基準: {structure_check['column_names']} ({structure_check['column_count']}列)")
            for error_file in structure_check["error_files"]:
                file_info = next((f for f in structure_check["files"] if f["path"] == error_file), None)
                if file_info:
                    print(f"  不一致: {error_file}")
                    print(f"    列名: {file_info['column_names']} ({file_info['column_count']}列)")
            raise ValueError("列構造が一致していません。処理を中止します。")

        print(f"[OK] 全ファイルの列構造が一致しています（{structure_check['column_count']}列）")
        print(f"  列名: {structure_check['column_names']}")

    # 出力ファイル名を自動生成
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_path = Path(file_paths[0]).parent
        output_path = base_path / "output"
        output_path.mkdir(exist_ok=True)
        output_file = output_path / f"multiple_files_translated_{timestamp}.csv"
    else:
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    result_data["output_file"] = str(output_file)

    # 全翻訳結果を収集
    all_results = []

    # 各ファイルを翻訳
    print(f"\n[INFO] {len(file_paths)}件のファイルを翻訳中...")
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] {Path(file_path).name}")
        try:
            # ファイルから単語を読み込み
            file = Path(file_path)
            words = []
            with open(file, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    if len(row) > column_index and row[column_index].strip():
                        words.append(row[column_index].strip())

            if not words:
                print(f"  [WARNING] 列インデックス {column_index} にデータが見つかりません")
                continue

            print(f"  読み込み: {len(words)}件")
            print(f"  翻訳実行中: 日本語 → {intermediate_lang} → 日本語")

            # 往復翻訳を実行
            results = round_trip_translate_batch(words, intermediate_lang)

            # ファイル名を追加
            for result in results:
                result["file_name"] = file.name
                all_results.append(result)

            # 統計情報を更新
            perfect_matches = sum(1 for r in results if r["is_perfect_match"])
            result_data["total_count"] += len(results)
            result_data["perfect_match_count"] += perfect_matches
            result_data["success_count"] += 1

            print(f"  完了: {len(results)}件翻訳、完全一致率 {perfect_matches / len(results) * 100:.1f}%")

        except Exception as e:
            result_data["error_count"] += 1
            error_info = {
                "file": file_path,
                "error": str(e)
            }
            result_data["errors"].append(error_info)
            print(f"  [ERROR] {e}")

    # 完全一致率を計算
    if result_data["total_count"] > 0:
        result_data["perfect_match_rate"] = (
            result_data["perfect_match_count"] / result_data["total_count"] * 100
        )

    # 1つのCSVファイルに出力
    if all_results:
        with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
            fieldnames = [
                "ファイル名",
                "元の日本語",
                "中間言語",
                "中間言語の翻訳",
                "日本語への逆翻訳",
                "完全一致"
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for result in all_results:
                writer.writerow({
                    "ファイル名": result["file_name"],
                    "元の日本語": result["original"],
                    "中間言語": result["intermediate_lang"],
                    "中間言語の翻訳": result["intermediate_text"],
                    "日本語への逆翻訳": result["back_translation"],
                    "完全一致": result["is_perfect_match"]
                })

        print(f"\n[INFO] 出力完了: {output_file}")
        print(f"  総件数: {result_data['total_count']}件")
        print(f"  完全一致率: {result_data['perfect_match_rate']:.1f}%")

    return result_data


def translate_from_multiple_excel(
    file_paths: List[str],
    output_file: Optional[str] = None,
    column_index: int = 0,
    intermediate_lang: str = "en",
    sheet_name: Optional[str] = None,
    check_structure: bool = True
) -> Dict:
    """
    複数のExcelファイルを一括で翻訳し、1つのCSVファイルにまとめて出力

    Args:
        file_paths (List[str]): 入力Excelファイルパスのリスト
        output_file (str): 出力CSVファイルのパス（省略時は自動生成）
        column_index (int): 翻訳する列のインデックス（0始まり、デフォルト: 0）
        intermediate_lang (str): 中間言語コード（デフォルト: "en"）
        sheet_name (str): シート名（省略時は最初のシート）
        check_structure (bool): 列構造チェックを実行するか（デフォルト: True）

    Returns:
        dict: 処理結果（translate_from_multiple_csvと同じ形式）

    Raises:
        ValueError: ファイルリストが空、または構造チェックで不一致が検出された場合
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxlがインストールされていません。'pip install openpyxl'を実行してください。")

    if not file_paths:
        raise ValueError("ファイルパスのリストが空です")

    result_data = {
        "structure_check": None,
        "output_file": None,
        "total_files": len(file_paths),
        "total_count": 0,
        "perfect_match_count": 0,
        "perfect_match_rate": 0.0,
        "success_count": 0,
        "error_count": 0,
        "errors": []
    }

    # 列構造チェック
    if check_structure:
        print("[INFO] 列構造チェック実行中...")
        structure_check = check_excel_structure(file_paths, sheet_name)
        result_data["structure_check"] = structure_check

        if not structure_check["is_valid"]:
            print("[ERROR] 列構造が一致していません")
            print(f"  基準: {structure_check['column_names']} ({structure_check['column_count']}列)")
            for error_file in structure_check["error_files"]:
                file_info = next((f for f in structure_check["files"] if f["path"] == error_file), None)
                if file_info:
                    print(f"  不一致: {error_file}")
                    print(f"    列名: {file_info['column_names']} ({file_info['column_count']}列)")
            raise ValueError("列構造が一致していません。処理を中止します。")

        print(f"[OK] 全ファイルの列構造が一致しています（{structure_check['column_count']}列）")
        print(f"  列名: {structure_check['column_names']}")

    # 出力ファイル名を自動生成
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_path = Path(file_paths[0]).parent
        output_path = base_path / "output"
        output_path.mkdir(exist_ok=True)
        output_file = output_path / f"multiple_files_translated_{timestamp}.csv"
    else:
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    result_data["output_file"] = str(output_file)

    # 全翻訳結果を収集
    all_results = []

    # 各ファイルを翻訳
    print(f"\n[INFO] {len(file_paths)}件のファイルを翻訳中...")
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] {Path(file_path).name}")
        try:
            # ファイルから単語を読み込み
            file = Path(file_path)
            wb = load_workbook(file, read_only=True, data_only=True)
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"シート '{sheet_name}' が見つかりません。利用可能なシート: {wb.sheetnames}")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            words = []
            for row in ws.iter_rows(values_only=True):
                if row and len(row) > column_index and row[column_index]:
                    word = str(row[column_index]).strip()
                    if word:
                        words.append(word)

            wb.close()

            if not words:
                print(f"  [WARNING] 列インデックス {column_index} にデータが見つかりません")
                continue

            print(f"  読み込み: {len(words)}件")
            print(f"  翻訳実行中: 日本語 → {intermediate_lang} → 日本語")

            # 往復翻訳を実行
            results = round_trip_translate_batch(words, intermediate_lang)

            # ファイル名を追加
            for result in results:
                result["file_name"] = file.name
                all_results.append(result)

            # 統計情報を更新
            perfect_matches = sum(1 for r in results if r["is_perfect_match"])
            result_data["total_count"] += len(results)
            result_data["perfect_match_count"] += perfect_matches
            result_data["success_count"] += 1

            print(f"  完了: {len(results)}件翻訳、完全一致率 {perfect_matches / len(results) * 100:.1f}%")

        except Exception as e:
            result_data["error_count"] += 1
            error_info = {
                "file": file_path,
                "error": str(e)
            }
            result_data["errors"].append(error_info)
            print(f"  [ERROR] {e}")

    # 完全一致率を計算
    if result_data["total_count"] > 0:
        result_data["perfect_match_rate"] = (
            result_data["perfect_match_count"] / result_data["total_count"] * 100
        )

    # 1つのCSVファイルに出力
    if all_results:
        with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
            fieldnames = [
                "ファイル名",
                "元の日本語",
                "中間言語",
                "中間言語の翻訳",
                "日本語への逆翻訳",
                "完全一致"
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for result in all_results:
                writer.writerow({
                    "ファイル名": result["file_name"],
                    "元の日本語": result["original"],
                    "中間言語": result["intermediate_lang"],
                    "中間言語の翻訳": result["intermediate_text"],
                    "日本語への逆翻訳": result["back_translation"],
                    "完全一致": result["is_perfect_match"]
                })

        print(f"\n[INFO] 出力完了: {output_file}")
        print(f"  総件数: {result_data['total_count']}件")
        print(f"  完全一致率: {result_data['perfect_match_rate']:.1f}%")

    return result_data


if __name__ == "__main__":
    print("=" * 70)
    print("一括翻訳スクリプト")
    print("=" * 70)
    print("\nこのスクリプトは直接実行せず、他のスクリプトからインポートして使用してください。")
    print("\n使用例:")
    print("  from batch_translator import translate_from_csv, translate_from_excel")
    print("  from batch_translator import translate_from_multiple_csv, translate_from_multiple_excel")
    print("  from batch_translator import check_csv_structure, check_excel_structure")
    print("")
    print("  # 単一ファイル")
    print("  result = translate_from_csv('input.csv', column_index=0, intermediate_lang='en')")
    print("")
    print("  # 複数ファイル（1つのCSVにまとめて出力、ファイル名列付き）")
    print("  files = ['file1.csv', 'file2.csv', 'file3.csv']")
    print("  result = translate_from_multiple_csv(files, column_index=0, intermediate_lang='en')")
    print("  print(f\"出力: {result['output_file']}\")")
    print("  print(f\"総件数: {result['total_count']}件\")")
    print("")
    print("  # 列構造チェックのみ")
    print("  check = check_csv_structure(files)")
    print("  print(f\"一致: {check['is_valid']}\")")
    print("=" * 70)
