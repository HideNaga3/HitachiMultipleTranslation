"""
テスト用のCSV/Excelファイルを作成するスクリプト
"""
import csv
from pathlib import Path
from openpyxl import Workbook


def create_test_csv():
    """
    テスト用CSVファイルを作成
    """
    test_dir = Path(__file__).parent.parent / "test_data"
    test_dir.mkdir(exist_ok=True)

    # テストデータ（日本語の単語リスト）
    test_words = [
        "平和",
        "希望",
        "未来",
        "勇気",
        "友情",
        "愛情",
        "自由",
        "正義",
        "幸福",
        "知恵",
        "感謝",
        "誠実",
        "優しさ",
        "強さ",
        "絆",
        "夢",
        "挑戦",
        "成長",
        "調和",
        "信頼"
    ]

    # CSV1: 単一列
    csv_file1 = test_dir / "test_words_single_column.csv"
    with open(csv_file1, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["日本語"])
        for word in test_words:
            writer.writerow([word])

    print(f"[OK] 作成完了: {csv_file1}")
    print(f"     データ件数: {len(test_words)}件")

    # CSV2: 複数列（列インデックス指定テスト用）
    csv_file2 = test_dir / "test_words_multi_column.csv"
    with open(csv_file2, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "カテゴリ", "日本語", "備考"])
        for i, word in enumerate(test_words, 1):
            category = "抽象概念" if i <= 10 else "価値観"
            writer.writerow([i, category, word, f"テスト{i}"])

    print(f"[OK] 作成完了: {csv_file2}")
    print(f"     データ件数: {len(test_words)}件")
    print(f"     列構成: ID, カテゴリ, 日本語(index=2), 備考")

    return csv_file1, csv_file2


def create_test_excel():
    """
    テスト用Excelファイルを作成
    """
    test_dir = Path(__file__).parent.parent / "test_data"
    test_dir.mkdir(exist_ok=True)

    # テストデータ（日本語の単語リスト）
    test_words = [
        "平和",
        "希望",
        "未来",
        "勇気",
        "友情",
        "愛情",
        "自由",
        "正義",
        "幸福",
        "知恵",
        "感謝",
        "誠実",
        "優しさ",
        "強さ",
        "絆",
        "夢",
        "挑戦",
        "成長",
        "調和",
        "信頼"
    ]

    # Excel1: 単一シート、単一列
    excel_file1 = test_dir / "test_words_single_column.xlsx"
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "単語リスト"
    ws1.append(["日本語"])
    for word in test_words:
        ws1.append([word])
    wb1.save(excel_file1)

    print(f"[OK] 作成完了: {excel_file1}")
    print(f"     データ件数: {len(test_words)}件")

    # Excel2: 単一シート、複数列（列インデックス指定テスト用）
    excel_file2 = test_dir / "test_words_multi_column.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "単語リスト"
    ws2.append(["ID", "カテゴリ", "日本語", "備考"])
    for i, word in enumerate(test_words, 1):
        category = "抽象概念" if i <= 10 else "価値観"
        ws2.append([i, category, word, f"テスト{i}"])
    wb2.save(excel_file2)

    print(f"[OK] 作成完了: {excel_file2}")
    print(f"     データ件数: {len(test_words)}件")
    print(f"     列構成: ID, カテゴリ, 日本語(index=2), 備考")

    # Excel3: 複数シート
    excel_file3 = test_dir / "test_words_multi_sheet.xlsx"
    wb3 = Workbook()

    # シート1: 抽象概念
    ws3_1 = wb3.active
    ws3_1.title = "抽象概念"
    ws3_1.append(["日本語"])
    for word in test_words[:10]:
        ws3_1.append([word])

    # シート2: 価値観
    ws3_2 = wb3.create_sheet("価値観")
    ws3_2.append(["日本語"])
    for word in test_words[10:]:
        ws3_2.append([word])

    wb3.save(excel_file3)

    print(f"[OK] 作成完了: {excel_file3}")
    print(f"     シート1「抽象概念」: {len(test_words[:10])}件")
    print(f"     シート2「価値観」: {len(test_words[10:])}件")

    return excel_file1, excel_file2, excel_file3


if __name__ == "__main__":
    print("=" * 70)
    print("テストファイル作成スクリプト")
    print("=" * 70)

    print("\n[1] CSVファイル作成中...")
    csv_files = create_test_csv()

    print("\n[2] Excelファイル作成中...")
    excel_files = create_test_excel()

    print("\n" + "=" * 70)
    print("[OK] すべてのテストファイルを作成しました")
    print("=" * 70)
    print("\n作成されたファイル:")
    print("  CSVファイル:")
    for f in csv_files:
        print(f"    - {f.name}")
    print("  Excelファイル:")
    for f in excel_files:
        print(f"    - {f.name}")
    print("\n保存先: test_data/")
    print("=" * 70)
