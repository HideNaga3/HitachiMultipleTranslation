"""
3シート比較Excel作成スクリプト（本番用 - 513行全データ版）
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from dotenv import load_dotenv
import requests
import html
import time
from difflib import SequenceMatcher

# .envファイルから環境変数を読み込み
load_dotenv()


def translate_with_api_key(texts, source_lang: str, target_lang: str, api_key: str):
    """Google Translate REST APIを使って翻訳"""
    base_url = "https://translation.googleapis.com/language/translate/v2"

    single_input = isinstance(texts, str)
    if single_input:
        texts = [texts]

    params = {
        "key": api_key,
        "q": texts,
        "source": source_lang,
        "target": target_lang,
        "format": "text"
    }

    response = requests.post(base_url, params=params)
    response.raise_for_status()

    result = response.json()
    translations = [html.unescape(t['translatedText']) for t in result['data']['translations']]

    return translations[0] if single_input else translations


def calculate_similarity(str1, str2):
    """2つの文字列の類似度を計算"""
    if pd.isna(str1) or pd.isna(str2):
        return 0.0
    if str1 == str2:
        return 100.0

    matcher = SequenceMatcher(None, str(str1), str(str2))
    return round(matcher.ratio() * 100, 1)


# 入力ファイル（本番用 - 全データ513行）
input_csv = r'C:\python_script\test_space\MitsubishiMultipleTranslation\core_files\for_import_インドネシア.csv'
output_excel = r'C:\python_script\test_space\MitsubishiMultipleTranslation\output\比較用_インドネシア.xlsx'

print('=' * 80)
print('3シート比較Excel作成（本番版 - 全513行）')
print('=' * 80)
print(f'実行日時: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
print()

# Google Translate API初期化
api_key = os.getenv('GOOGLE_API_KEY')
if not api_key:
    raise ValueError("Google API キーが設定されていません")

print('✓ Google Translate API: 初期化成功')
print()

# CSV読み込み
df = pd.read_csv(input_csv, encoding='utf-8-sig')

print(f'✓ 入力CSV: {os.path.basename(input_csv)}')
print(f'  行数: {len(df)}行')
print(f'  列数: {len(df.columns)}列')
print()

# Google Translate APIの言語コードマッピング
api_lang_codes = {
    'en': 'en', 'fil-PH': 'tl', 'pt': 'pt', 'es': 'es', 'pt-BR': 'pt',
    'zh': 'zh-CN', 'ko': 'ko', 'fr': 'fr', 'hi': 'hi', 'th': 'th',
    'vi': 'vi', 'my': 'my', 'ne': 'ne', 'bn': 'bn', 'id': 'id',
    'ta': 'ta', 'si': 'si', 'mn': 'mn', 'ar': 'ar', 'fa': 'fa',
    'tr': 'tr', 'ru': 'ru', 'ur': 'ur', 'km': 'km', 'lo': 'lo',
    'ms': 'ms', 'de': 'de', 'hu': 'hu', 'cs': 'cs', 'pl': 'pl',
    'nl': 'nl', 'da': 'da', 'fi': 'fi', 'sv': 'sv', 'lb': 'lb',
    'af': 'af', 'fr-CA': 'fr',
}

# 翻訳対象の列
columns_to_translate = [col for col in df.columns if col != 'ja' and col in api_lang_codes]

# 出力用データフレーム
df_retranslated = df.copy()
comparison_data = []

print('=' * 80)
print('逆翻訳処理開始')
print('=' * 80)
print(f'推定処理時間: 約10分（1,255件のユニーク翻訳）')
print()

# 各言語列を逆翻訳
for col_idx_loop, col_code in enumerate(columns_to_translate, 1):
    if col_code not in df.columns:
        continue

    print(f'[{col_idx_loop}/{len(columns_to_translate)}] 処理中: {col_code}')

    # 列のインデックス番号
    col_idx = df.columns.get_loc(col_code)
    col_letter = get_column_letter(col_idx + 1)

    # 空欄でない翻訳データを収集
    non_empty_texts = []
    for text in df[col_code]:
        if pd.isna(text):
            continue
        text_str = str(text).strip()
        if text_str == '':
            continue
        non_empty_texts.append(text_str)

    unique_texts = list(set(non_empty_texts))

    print(f'  空欄でないセル: {len(non_empty_texts)}個')
    print(f'  ユニークな翻訳: {len(unique_texts)}件')

    if len(unique_texts) == 0:
        print(f'  → スキップ')
        print()
        continue

    # 逆翻訳
    reverse_translations = {}
    translated_count = 0

    for row_idx, text in enumerate(df[col_code], start=2):
        if pd.isna(text):
            continue
        text_str = str(text).strip()
        if text_str == '':
            continue

        if text_str in reverse_translations:
            continue

        try:
            source_lang = api_lang_codes[col_code]
            translated = translate_with_api_key(text_str, source_lang=source_lang, target_lang='ja', api_key=api_key)
            reverse_translations[text_str] = translated
            translated_count += 1

            # 進捗表示（50件ごと）
            if translated_count % 50 == 0:
                print(f'    進捗: {translated_count}/{len(unique_texts)}件')

            time.sleep(0.1)

        except Exception as e:
            print(f'  エラー（行{row_idx}）: {e}')
            reverse_translations[text_str] = ''

    # 結果を適用
    def apply_retranslation(x):
        if pd.isna(x):
            return x
        x_str = str(x).strip()
        if x_str == '':
            return x
        return reverse_translations.get(x_str, x)

    df_retranslated[col_code] = df[col_code].apply(apply_retranslation)

    print(f'  ✓ 完了: {len(reverse_translations)}件')
    print()

    # 比較データを作成
    for row_idx, (ja_word, translation, retranslation) in enumerate(
        zip(df['ja'], df[col_code], df_retranslated[col_code]), start=2
    ):
        if pd.isna(translation) or str(translation).strip() == '':
            continue

        cell_address = f'{col_letter}{row_idx}'
        is_match = (ja_word == retranslation)
        similarity = calculate_similarity(ja_word, retranslation)
        classification = '完全一致' if is_match else '不一致'

        comparison_data.append({
            'アドレス': cell_address,
            '言語': col_code,
            '単語': ja_word,
            '再翻訳': retranslation,
            '翻訳': translation,
            '一致': is_match,
            '類似度_difflib': f'{similarity}%',
            '分類': classification
        })

print('=' * 80)
print('逆翻訳処理完了')
print('=' * 80)
print()

# 比較データフレーム
df_comparison = pd.DataFrame(comparison_data)

print(f'✓ 比較データ: {len(df_comparison)}行')
print()

# Excelファイル作成
print('Excelファイル作成中...')
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='翻訳', index=False, header=True)
    df_retranslated.to_excel(writer, sheet_name='再翻訳', index=False, header=True)
    df_comparison.to_excel(writer, sheet_name='比較', index=False, header=True)

print(f'  翻訳シート: {len(df)+1}行')
print(f'  再翻訳シート: {len(df_retranslated)+1}行')
print(f'  比較シート: {len(df_comparison)+1}行')
print()

# スタイル適用
print('スタイル適用中...')
wb = openpyxl.load_workbook(output_excel)

ws_translation = wb['翻訳']
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws_translation[1]:
    cell.fill = blue_fill
    cell.font = Font(bold=True)

ws_retranslation = wb['再翻訳']
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
for cell in ws_retranslation[1]:
    cell.fill = green_fill
    cell.font = Font(bold=True)

ws_comparison = wb['比較']
orange_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
for cell in ws_comparison[1]:
    cell.fill = orange_fill
    cell.font = Font(bold=True)

# 列幅を自動調整
for ws in [ws_translation, ws_retranslation, ws_comparison]:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_excel)
print()

file_size_kb = os.path.getsize(output_excel) / 1024
print(f'✓ 出力: {output_excel}')
print(f'  ファイルサイズ: {file_size_kb:.2f} KB')
print()

# 統計情報
total_comparisons = len(df_comparison)
perfect_matches = len(df_comparison[df_comparison['一致'] == True])
match_rate = (perfect_matches / total_comparisons * 100) if total_comparisons > 0 else 0

print('=' * 80)
print('統計情報')
print('=' * 80)
print(f'比較総数: {total_comparisons}件')
print(f'完全一致: {perfect_matches}件')
print(f'不一致: {total_comparisons - perfect_matches}件')
print(f'一致率: {match_rate:.1f}%')
print()

print('=' * 80)
print('完了')
print('=' * 80)
