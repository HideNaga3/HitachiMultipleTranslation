"""
CSV → 逆翻訳Excel作成関数
UTF-8 BOM形式のCSVを入力し、各言語を日本語に逆翻訳したExcelを出力
"""
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
from dotenv import load_dotenv
import requests
import html
from typing import Optional, Dict, List

# .envファイルから環境変数を読み込み
load_dotenv()


def translate_with_api_key(texts, source_lang: str, target_lang: str, api_key: str):
    """
    Google Translate REST APIを使って翻訳（複数テキスト対応）

    Args:
        texts: 翻訳するテキスト（文字列またはリスト）
        source_lang: ソース言語コード
        target_lang: ターゲット言語コード
        api_key: Google API キー

    Returns:
        翻訳結果（文字列またはリスト）
    """
    base_url = "https://translation.googleapis.com/language/translate/v2"

    # 単一文字列の場合はリストに変換
    single_input = isinstance(texts, str)
    if single_input:
        texts = [texts]

    params = {
        "key": api_key,
        "q": texts,
        "source": source_lang,
        "target": target_lang,
        "format": "text"
    }

    response = requests.post(base_url, params=params)
    response.raise_for_status()

    result = response.json()
    # HTMLエスケープをデコード
    translations = [html.unescape(t['translatedText']) for t in result['data']['translations']]

    # 単一入力の場合は文字列を返す
    return translations[0] if single_input else translations


def create_reverse_translation_excel(
    input_csv: str,
    output_excel: Optional[str] = None,
    columns: Optional[List[str]] = None,
    verbose: bool = True
) -> Dict:
    """
    CSV → 逆翻訳Excel作成

    Args:
        input_csv (str): 入力CSVファイルパス（UTF-8 BOM）
        output_excel (str): 出力Excelファイルパス（省略時は自動生成）
        columns (List[str]): 逆翻訳する列名のリスト（省略時は全言語列）
        verbose (bool): 進捗表示（デフォルト: True）

    Returns:
        dict: 処理結果
            {
                "input_csv": 入力ファイルパス,
                "output_excel": 出力ファイルパス,
                "row_count": 行数,
                "translated_columns": 翻訳した列のリスト
            }

    Raises:
        FileNotFoundError: 入力ファイルが見つからない
        ValueError: API認証エラー
    """
    input_path = Path(input_csv)
    if not input_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_csv}")

    # 出力ファイル名を自動生成
    if output_excel is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = input_path.parent / "output"
        output_dir.mkdir(exist_ok=True)
        output_excel = output_dir / f"逆翻訳_検証結果_{timestamp}.xlsx"
    else:
        output_excel = Path(output_excel)
        output_excel.parent.mkdir(parents=True, exist_ok=True)

    if verbose:
        print("=" * 80)
        print("Google Translate API による逆翻訳（翻訳→日本語）")
        print("=" * 80)
        print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()

    # Google Translate API初期化
    api_key = os.getenv('GOOGLE_API_KEY')
    if not api_key:
        raise ValueError(
            "Google API キーが設定されていません。\n"
            ".envファイルに GOOGLE_API_KEY=your_api_key_here を設定してください。"
        )

    if verbose:
        print("Google Translate API: 初期化成功")
        print("認証方法: APIキー（.envファイルから読み込み）")
        print()

    # CSV読み込み
    df = pd.read_csv(input_path, encoding='utf-8-sig')

    if verbose:
        print(f"入力CSV: {input_path.name}")
        print(f"行数: {len(df)}行")
        print(f"列数: {len(df.columns)}列")
        print()

    # 言語コードと列名のマッピング
    lang_mapping = {
        'en': '英語',
        'fil-PH': 'タガログ語',
        'zh': '中国語',
        'th': 'タイ語',
        'vi': 'ベトナム語',
        'my': 'ミャンマー語',
        'id': 'インドネシア語',
        'km': 'カンボジア語'
    }

    # Google Translate APIの言語コードマッピング
    api_lang_codes = {
        'en': 'en',
        'fil-PH': 'tl',  # Tagalog
        'zh': 'zh-CN',   # Chinese (Simplified)
        'th': 'th',
        'vi': 'vi',
        'my': 'my',
        'id': 'id',
        'km': 'km'
    }

    # 翻訳対象の列を決定
    if columns is None:
        columns = list(lang_mapping.keys())

    # 出力用データフレーム（元のデータをコピー）
    df_output = df.copy()

    if verbose:
        print("翻訳対象の列:")
        for col_code in columns:
            if col_code in lang_mapping:
                print(f"  {col_code:10s} → {lang_mapping[col_code]}")
        print()

    # 各言語列を日本語に逆翻訳
    if verbose:
        print("=" * 80)
        print("逆翻訳処理開始")
        print("=" * 80)
        print()

    translated_columns = []

    for col_code in columns:
        if col_code not in df.columns:
            if verbose:
                print(f"警告: 列 '{col_code}' が見つかりません。スキップします。")
            continue

        lang_name = lang_mapping.get(col_code, col_code)

        if verbose:
            print(f"処理中: {lang_name} ({col_code})")

        # 翻訳対象のテキストを取得（空欄を除く）
        translations = df[col_code].dropna().tolist()

        # 空文字列を除外
        translations = [t for t in translations if isinstance(t, str) and t.strip() != '']

        if verbose:
            print(f"  翻訳対象: {len(translations)}件")

        if len(translations) == 0:
            if verbose:
                print(f"  スキップ: 翻訳対象なし")
                print()
            continue

        # 逆翻訳結果を格納
        reverse_translations = {}

        # バッチ処理（API制限対策）
        batch_size = 20
        total_batches = (len(translations) + batch_size - 1) // batch_size

        for batch_idx in range(total_batches):
            start_idx = batch_idx * batch_size
            end_idx = min((batch_idx + 1) * batch_size, len(translations))
            batch = translations[start_idx:end_idx]

            if verbose:
                print(f"  バッチ {batch_idx + 1}/{total_batches}: {len(batch)}件")

            try:
                # Google Translate API で翻訳
                source_lang = api_lang_codes[col_code]

                # 未翻訳のテキストのみを抽出
                texts_to_translate = [t for t in batch if t not in reverse_translations]

                if len(texts_to_translate) > 0:
                    # REST APIで翻訳（APIキー使用・バッチ処理）
                    translated_texts = translate_with_api_key(
                        texts_to_translate,
                        source_lang=source_lang,
                        target_lang='ja',
                        api_key=api_key
                    )
                    # 結果を辞書に格納
                    for original, translated in zip(texts_to_translate, translated_texts):
                        reverse_translations[original] = translated

                # API制限対策（1秒待機）
                if batch_idx < total_batches - 1:
                    time.sleep(1)

            except Exception as e:
                if verbose:
                    print(f"  エラー: {e}")
                    print(f"  バッチ {batch_idx + 1} をスキップします")

        # 逆翻訳結果をデータフレームに適用
        df_output[col_code] = df[col_code].apply(
            lambda x: reverse_translations.get(x, x) if pd.notna(x) and x != '' else x
        )

        translated_columns.append(col_code)

        if verbose:
            print(f"  完了: {len(reverse_translations)}件の逆翻訳")
            print()

    if verbose:
        print("=" * 80)
        print("逆翻訳処理完了")
        print("=" * 80)
        print()

    # Excelファイルに保存
    if verbose:
        print("Excelファイル作成中...")

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # inputシート: 元のデータ
        df.to_excel(writer, sheet_name='input', index=False)

        # outputシート: 逆翻訳結果
        df_output.to_excel(writer, sheet_name='output', index=False)

        if verbose:
            print(f"  inputシート: {len(df)}行 x {len(df.columns)}列")
            print(f"  outputシート: {len(df_output)}行 x {len(df_output.columns)}列")

    # スタイル適用（見やすくするため）
    wb = openpyxl.load_workbook(output_excel)

    # inputシートのヘッダーを青色に
    ws_input = wb['input']
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    for cell in ws_input[1]:
        cell.fill = blue_fill
        cell.font = Font(bold=True)

    # outputシートのヘッダーを緑色に
    ws_output = wb['output']
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    for cell in ws_output[1]:
        cell.fill = green_fill
        cell.font = Font(bold=True)

    # 列幅を自動調整
    for ws in [ws_input, ws_output]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)

    if verbose:
        print()
        print(f"保存: {output_excel}")
        print()
        print("=" * 80)
        print("完了")
        print("=" * 80)

    return {
        "input_csv": str(input_path),
        "output_excel": str(output_excel),
        "row_count": len(df),
        "translated_columns": translated_columns
    }


if __name__ == "__main__":
    # 使用例
    print("=" * 70)
    print("逆翻訳Excel作成スクリプト")
    print("=" * 70)
    print("\n使用例:")
    print("  from scripts.utils.reverse_translate_excel import create_reverse_translation_excel")
    print("")
    print("  result = create_reverse_translation_excel(")
    print("      input_csv='output/全言語統合_テンプレート_インポート用.csv',")
    print("      output_excel='output/逆翻訳_検証結果.xlsx'")
    print("  )")
    print("")
    print("  print(f\"出力: {result['output_excel']}\")")
    print("  print(f\"行数: {result['row_count']}\")")
    print("=" * 70)
