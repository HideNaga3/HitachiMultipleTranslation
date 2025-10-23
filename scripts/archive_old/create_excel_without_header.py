"""
ヘッダーなしのExcelファイルを作成
"""
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

output_dir = r'C:\python_script\test_space\MitsubishiMultipleTranslation\output'
deliverables_dir = r'C:\python_script\test_space\MitsubishiMultipleTranslation\成果物_20251023'

print("=" * 80)
print("ヘッダーなしExcelファイルの作成")
print("=" * 80)
print()

# テンプレートCSV読み込み
input_csv = os.path.join(output_dir, '全言語統合_テンプレート_インポート用.csv')
df = pd.read_csv(input_csv, encoding='utf-8-sig')

print(f"入力CSV: {os.path.basename(input_csv)}")
print(f"行数: {len(df)}行（ヘッダー含む）")
print(f"列数: {len(df.columns)}列")
print()

# 逆翻訳結果読み込み
excel_file = os.path.join(output_dir, '逆翻訳_検証結果.xlsx')
df_input = pd.read_excel(excel_file, sheet_name='input')
df_output = pd.read_excel(excel_file, sheet_name='output')

print(f"入力Excel: {os.path.basename(excel_file)}")
print(f"  inputシート: {len(df_input)}行")
print(f"  outputシート: {len(df_output)}行")
print()

# ヘッダーなしExcelファイル作成
output_excel = os.path.join(deliverables_dir, '02_逆翻訳_検証結果.xlsx')

print("ヘッダーなしExcelファイル作成中...")

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    # inputシート（ヘッダーなし）
    df_input.to_excel(writer, sheet_name='input', index=False, header=False)
    # outputシート（ヘッダーなし）
    df_output.to_excel(writer, sheet_name='output', index=False, header=False)

print(f"  inputシート: {len(df_input)}行 x {len(df_input.columns)}列（ヘッダーなし）")
print(f"  outputシート: {len(df_output)}行 x {len(df_output.columns)}列（ヘッダーなし）")
print()

# スタイル適用
wb = openpyxl.load_workbook(output_excel)

# inputシートの1行目を青色に（データの先頭行）
ws_input = wb['input']
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws_input[1]:
    cell.fill = blue_fill
    cell.font = Font(bold=True)

# outputシートの1行目を緑色に（データの先頭行）
ws_output = wb['output']
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
for cell in ws_output[1]:
    cell.fill = green_fill
    cell.font = Font(bold=True)

# 列幅を自動調整
for ws in [ws_input, ws_output]:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_excel)

print(f"保存: {output_excel}")
print()
print("=" * 80)
print("完了")
print("=" * 80)
print()
print("注意:")
print("  - CSVファイルはヘッダーあり（そのまま）")
print("  - Excelファイルはヘッダーなし（更新済み）")
print("  - Excel1行目は元のヘッダー行だったデータ（色付き）")
print()
