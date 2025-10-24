"""
3シート比較Excel作成スクリプト
CSVを入力 → 逆翻訳 → 3シートExcel出力

シート構成:
1. 翻訳: 元のデータ
2. 再翻訳: 逆翻訳結果
3. 比較: セルごとの比較詳細
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from dotenv import load_dotenv
import requests
import html
import time
from difflib import SequenceMatcher

# .envファイルから環境変数を読み込み
load_dotenv()


def translate_with_api_key(texts, source_lang: str, target_lang: str, api_key: str):
    """
    Google Translate REST APIを使って翻訳（複数テキスト対応）
    """
    base_url = "https://translation.googleapis.com/language/translate/v2"

    # 単一文字列の場合はリストに変換
    single_input = isinstance(texts, str)
    if single_input:
        texts = [texts]

    params = {
        "key": api_key,
        "q": texts,
        "source": source_lang,
        "target": target_lang,
        "format": "text"
    }

    response = requests.post(base_url, params=params)
    response.raise_for_status()

    result = response.json()
    # HTMLエスケープをデコード
    translations = [html.unescape(t['translatedText']) for t in result['data']['translations']]

    # 単一入力の場合は文字列を返す
    return translations[0] if single_input else translations


def calculate_similarity(str1, str2):
    """
    2つの文字列の類似度を計算（difflib使用）
    """
    if pd.isna(str1) or pd.isna(str2):
        return 0.0
    if str1 == str2:
        return 100.0

    matcher = SequenceMatcher(None, str(str1), str(str2))
    return round(matcher.ratio() * 100, 1)


def create_3sheet_comparison_excel(input_csv: str, output_excel: str = None):
    """
    CSVから3シート比較Excelを作成

    Args:
        input_csv: 入力CSVファイルパス
        output_excel: 出力Excelファイルパス（省略時は自動生成）

    Returns:
        出力Excelファイルパス
    """
    print('=' * 80)
    print('3シート比較Excel作成')
    print('=' * 80)
    print(f'実行日時: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print()

    # Google Translate API初期化
    api_key = os.getenv('GOOGLE_API_KEY')
    if not api_key:
        raise ValueError(
            "Google API キーが設定されていません。\n"
            ".envファイルに GOOGLE_API_KEY=your_api_key_here を設定してください。"
        )

    print('Google Translate API: 初期化成功')
    print()

    # CSV読み込み
    df = pd.read_csv(input_csv, encoding='utf-8-sig')

    print(f'入力CSV: {os.path.basename(input_csv)}')
    print(f'  行数: {len(df)}行')
    print(f'  列数: {len(df.columns)}列')
    print()

    # 言語コードと列名のマッピング
    lang_mapping = {
        'en': '英語',
        'fil-PH': 'タガログ語',
        'pt': 'ポルトガル語',
        'es': 'スペイン語',
        'pt-BR': 'ポルトガル語(ブラジル)',
        'zh': '中国語',
        'ko': '韓国語',
        'fr': 'フランス語',
        'hi': 'ヒンディー語',
        'th': 'タイ語',
        'vi': 'ベトナム語',
        'my': 'ミャンマー語',
        'ne': 'ネパール語',
        'bn': 'ベンガル語',
        'id': 'インドネシア語',
        'ta': 'タミル語',
        'si': 'シンハラ語',
        'mn': 'モンゴル語',
        'ar': 'アラビア語',
        'fa': 'ペルシア語',
        'tr': 'トルコ語',
        'ru': 'ロシア語',
        'ur': 'ウルドゥー語',
        'km': 'カンボジア語',
        'lo': 'ラオス語',
        'ms': 'マレー語',
        'de': 'ドイツ語',
        'hu': 'ハンガリー語',
        'cs': 'チェコ語',
        'pl': 'ポーランド語',
        'nl': 'オランダ語',
        'da': 'デンマーク語',
        'fi': 'フィンランド語',
        'sv': 'スウェーデン語',
        'lb': 'ルクセンブルク語',
        'af': 'アフリカーンス語',
        'fr-CA': 'フランス語(カナダ)',
    }

    # Google Translate APIの言語コードマッピング
    api_lang_codes = {
        'en': 'en',
        'fil-PH': 'tl',
        'pt': 'pt',
        'es': 'es',
        'pt-BR': 'pt',
        'zh': 'zh-CN',
        'ko': 'ko',
        'fr': 'fr',
        'hi': 'hi',
        'th': 'th',
        'vi': 'vi',
        'my': 'my',
        'ne': 'ne',
        'bn': 'bn',
        'id': 'id',
        'ta': 'ta',
        'si': 'si',
        'mn': 'mn',
        'ar': 'ar',
        'fa': 'fa',
        'tr': 'tr',
        'ru': 'ru',
        'ur': 'ur',
        'km': 'km',
        'lo': 'lo',
        'ms': 'ms',
        'de': 'de',
        'hu': 'hu',
        'cs': 'cs',
        'pl': 'pl',
        'nl': 'nl',
        'da': 'da',
        'fi': 'fi',
        'sv': 'sv',
        'lb': 'lb',
        'af': 'af',
        'fr-CA': 'fr',
    }

    # 翻訳対象の列を決定（jaを除く全列）
    columns_to_translate = [col for col in df.columns if col != 'ja' and col in api_lang_codes]

    print('翻訳対象の列:')
    for col_code in columns_to_translate:
        lang_name = lang_mapping.get(col_code, col_code)
        print(f'  {col_code:10s} → {lang_name}')
    print()

    # 出力用データフレーム（逆翻訳結果）
    df_retranslated = df.copy()

    # 比較データを格納するリスト
    comparison_data = []

    print('=' * 80)
    print('逆翻訳処理開始')
    print('=' * 80)
    print()

    # 各言語列を日本語に逆翻訳
    for col_idx_loop, col_code in enumerate(columns_to_translate, 1):
        if col_code not in df.columns:
            continue

        lang_name = lang_mapping.get(col_code, col_code)

        print(f'[{col_idx_loop}/{len(columns_to_translate)}] 処理中: {lang_name} ({col_code})')

        # 列のインデックス番号を取得
        col_idx = df.columns.get_loc(col_code)
        col_letter = get_column_letter(col_idx + 1)

        # 空欄でない翻訳データを収集
        non_empty_texts = []
        for text in df[col_code]:
            # 空欄チェック（NaN、空文字、空白のみを除外）
            if pd.isna(text):
                continue
            text_str = str(text).strip()
            if text_str == '':
                continue
            non_empty_texts.append(text_str)

        # 重複を除いたユニークなテキスト数
        unique_texts = list(set(non_empty_texts))

        print(f'  データ行数: {len(df)}行')
        print(f'  空欄でないセル: {len(non_empty_texts)}個')
        print(f'  ユニークな翻訳: {len(unique_texts)}件（これを逆翻訳します）')

        # 翻訳対象がない場合はスキップ
        if len(unique_texts) == 0:
            print(f'  スキップ: 翻訳対象なし')
            print()
            continue

        # 翻訳対象のテキストを取得（空欄を除く）
        reverse_translations = {}

        # 行ごとに処理
        translated_count = 0
        for row_idx, text in enumerate(df[col_code], start=2):  # Excelは1行目がヘッダーなので2から
            # 空欄チェック（厳格に）
            if pd.isna(text):
                continue
            text_str = str(text).strip()
            if text_str == '':
                continue

            # すでに翻訳済みならスキップ
            if text_str in reverse_translations:
                continue

            try:
                # Google Translate APIで逆翻訳
                source_lang = api_lang_codes[col_code]
                translated = translate_with_api_key(
                    text_str,
                    source_lang=source_lang,
                    target_lang='ja',
                    api_key=api_key
                )
                reverse_translations[text_str] = translated
                translated_count += 1

                # 進捗表示（10件ごと）
                if translated_count % 10 == 0:
                    print(f'    進捗: {translated_count}/{len(unique_texts)}件')

                # API制限対策
                time.sleep(0.1)

            except Exception as e:
                print(f'  エラー（行{row_idx}）: {e}')
                reverse_translations[text_str] = ''

        # 逆翻訳結果をデータフレームに適用（空欄はそのまま）
        def apply_retranslation(x):
            if pd.isna(x):
                return x
            x_str = str(x).strip()
            if x_str == '':
                return x
            return reverse_translations.get(x_str, x)

        df_retranslated[col_code] = df[col_code].apply(apply_retranslation)

        print(f'  ✓ 完了: {len(reverse_translations)}件の逆翻訳')
        print()

        # 比較データを作成
        for row_idx, (ja_word, translation, retranslation) in enumerate(
            zip(df['ja'], df[col_code], df_retranslated[col_code]), start=2
        ):
            # 空欄スキップ
            if pd.isna(translation) or str(translation).strip() == '':
                continue

            # セルアドレス
            cell_address = f'{col_letter}{row_idx}'

            # 一致判定
            is_match = (ja_word == retranslation)

            # 類似度計算
            similarity = calculate_similarity(ja_word, retranslation)

            # 分類
            if is_match:
                classification = '完全一致'
            else:
                classification = '不一致'

            comparison_data.append({
                'アドレス': cell_address,
                '言語': col_code,
                '単語': ja_word,
                '再翻訳': retranslation,
                '翻訳': translation,
                '一致': is_match,
                '類似度_difflib': f'{similarity}%',
                '分類': classification
            })

    print('=' * 80)
    print('逆翻訳処理完了')
    print('=' * 80)
    print()

    # 比較データフレーム作成
    df_comparison = pd.DataFrame(comparison_data)

    print(f'比較データ: {len(df_comparison)}行')
    print()

    # 出力ファイル名決定
    if output_excel is None:
        output_dir = os.path.dirname(input_csv)
        if output_dir == '':
            output_dir = '.'
        output_dir = os.path.join(os.path.dirname(output_dir), 'output')
        os.makedirs(output_dir, exist_ok=True)

        # ファイル名から言語名を抽出
        basename = os.path.basename(input_csv)
        if 'インドネシア' in basename:
            lang_name = 'インドネシア'
        else:
            lang_name = basename.replace('for_import_', '').replace('.csv', '')

        output_excel = os.path.join(output_dir, f'比較用_{lang_name}.xlsx')

    # Excelファイル作成
    print('Excelファイル作成中...')
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # シート1: 翻訳（元のデータ）
        df.to_excel(writer, sheet_name='翻訳', index=False, header=True)

        # シート2: 再翻訳（逆翻訳結果）
        df_retranslated.to_excel(writer, sheet_name='再翻訳', index=False, header=True)

        # シート3: 比較
        df_comparison.to_excel(writer, sheet_name='比較', index=False, header=True)

    print(f'  翻訳シート: {len(df)+1}行（ヘッダー1行 + データ{len(df)}行）')
    print(f'  再翻訳シート: {len(df_retranslated)+1}行（ヘッダー1行 + データ{len(df_retranslated)}行）')
    print(f'  比較シート: {len(df_comparison)+1}行（ヘッダー1行 + データ{len(df_comparison)}行）')
    print()

    # スタイル適用
    print('スタイル適用中...')
    wb = openpyxl.load_workbook(output_excel)

    # 翻訳シートのヘッダーを青色に
    ws_translation = wb['翻訳']
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    for cell in ws_translation[1]:
        cell.fill = blue_fill
        cell.font = Font(bold=True)

    # 再翻訳シートのヘッダーを緑色に
    ws_retranslation = wb['再翻訳']
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    for cell in ws_retranslation[1]:
        cell.fill = green_fill
        cell.font = Font(bold=True)

    # 比較シートのヘッダーをオレンジ色に
    ws_comparison = wb['比較']
    orange_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    for cell in ws_comparison[1]:
        cell.fill = orange_fill
        cell.font = Font(bold=True)

    # 列幅を自動調整
    for ws in [ws_translation, ws_retranslation, ws_comparison]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)
    print()

    file_size_kb = os.path.getsize(output_excel) / 1024
    print(f'出力: {output_excel}')
    print(f'ファイルサイズ: {file_size_kb:.2f} KB')
    print()

    # 統計情報
    total_comparisons = len(df_comparison)
    perfect_matches = len(df_comparison[df_comparison['一致'] == True])
    match_rate = (perfect_matches / total_comparisons * 100) if total_comparisons > 0 else 0

    print('=' * 80)
    print('統計情報')
    print('=' * 80)
    print(f'比較総数: {total_comparisons}件')
    print(f'完全一致: {perfect_matches}件')
    print(f'不一致: {total_comparisons - perfect_matches}件')
    print(f'一致率: {match_rate:.1f}%')
    print()

    print('=' * 80)
    print('完了')
    print('=' * 80)

    return output_excel


if __name__ == "__main__":
    # 使用例
    input_csv = r'C:\python_script\test_space\MitsubishiMultipleTranslation\core_files\for_import_インドネシア.csv'

    output_excel = create_3sheet_comparison_excel(input_csv)

    print()
    print(f'作成されたファイル: {output_excel}')
