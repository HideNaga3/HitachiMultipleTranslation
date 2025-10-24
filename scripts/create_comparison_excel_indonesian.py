"""
比較用Excel作成スクリプト（インドネシア語用）
for_import_インドネシア.csv → 比較用_インドネシア.xlsx
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os

def create_comparison_excel():
    # 入力ファイル
    input_csv = r'C:\python_script\test_space\MitsubishiMultipleTranslation\core_files\for_import_インドネシア.csv'
    output_excel = r'C:\python_script\test_space\MitsubishiMultipleTranslation\output\比較用_インドネシア.xlsx'

    print('=' * 80)
    print('比較用Excel作成')
    print('=' * 80)
    print()

    # CSV読み込み
    df = pd.read_csv(input_csv, encoding='utf-8-sig')

    print(f'入力CSV: {os.path.basename(input_csv)}')
    print(f'  行数: {len(df)}行')
    print(f'  列数: {len(df.columns)}列')
    print()

    # Excelに保存
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='翻訳', index=False, header=True)

    print(f'シート作成: 翻訳シート ({len(df)+1}行 = ヘッダー1行 + データ{len(df)}行)')
    print()

    # スタイル適用
    print('スタイル適用中...')
    wb = openpyxl.load_workbook(output_excel)
    ws = wb['翻訳']

    # ヘッダーを青色に
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = Font(bold=True)

    # 列幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)
    print()

    file_size_kb = os.path.getsize(output_excel) / 1024
    print(f'出力: {output_excel}')
    print(f'ファイルサイズ: {file_size_kb:.2f} KB')
    print()
    print('=' * 80)
    print('完了')
    print('=' * 80)

    return output_excel


if __name__ == "__main__":
    create_comparison_excel()
